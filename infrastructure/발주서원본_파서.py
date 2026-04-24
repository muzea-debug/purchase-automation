"""
발주서원본_파서 - 통합발주서원본/*.xlsx → 바코드별 상품_정보 dict
"""
from pathlib import Path
import openpyxl
from domain.entities.상품_정보 import 상품_정보


class 발주서원본_파서:

    def __init__(self, 원본_디렉토리: Path):
        self._디렉토리 = 원본_디렉토리
        self._캐시: dict[str, 상품_정보] = {}
        self._캐시_mtime: dict[str, float] = {}  # 파일명 → 마지막 mtime

    def 카탈로그_로드(self) -> dict[str, 상품_정보]:
        """바코드 → 상품_정보 dict 반환. 파일 변경 없으면 캐시 재사용."""
        xlsx_파일들 = sorted(
            f for f in self._디렉토리.glob("*.xlsx")
            if not f.name.startswith("~$")
        )

        # 변경된 파일만 재파싱
        현재_키 = {f.name for f in xlsx_파일들}
        이전_키 = set(self._캐시_mtime.keys())

        변경됨 = False
        for f in xlsx_파일들:
            mtime = f.stat().st_mtime
            if self._캐시_mtime.get(f.name) != mtime:
                변경됨 = True
                try:
                    # 이 파일의 기존 캐시 항목 제거
                    self._캐시 = {
                        bc: info for bc, info in self._캐시.items()
                        if info.원본파일.name != f.name
                    }
                    self._파일_파싱(f, self._캐시)
                    self._캐시_mtime[f.name] = mtime
                    print(f"[발주서원본_파서] 파싱: {f.name}")
                except Exception as e:
                    print(f"[발주서원본_파서] 파싱 실패 {f.name}: {e}")

        # 삭제된 파일 캐시 정리
        삭제된 = 이전_키 - 현재_키
        if 삭제된:
            변경됨 = True
            for name in 삭제된:
                del self._캐시_mtime[name]
            self._캐시 = {
                bc: info for bc, info in self._캐시.items()
                if info.원본파일.name not in 삭제된
            }

        if not 변경됨:
            print(f"[발주서원본_파서] 캐시 사용 ({len(self._캐시)}개 상품)")

        return self._캐시

    def _파일_파싱(self, 파일: Path, 카탈로그: dict) -> None:
        제조사, 브랜드 = self._파일명_파싱(파일.stem)
        wb = openpyxl.load_workbook(파일, read_only=True, data_only=True)
        ws = wb.active
        계좌id = str(ws["J5"].value or "").strip()
        for row in ws.iter_rows(min_row=7, values_only=True):
            if len(row) < 12:
                break
            번호_val = row[1]   # B열
            if not 번호_val:
                break
            바코드_val = row[8]  # I열 - "한국상품명_바코드번호" 또는 바코드 단독
            if not 바코드_val:
                continue
            바코드_raw = str(바코드_val).strip()
            if not 바코드_raw:
                continue
            바코드_str, 한국_이름 = self._바코드_분리(바코드_raw)
            if not 바코드_str:
                continue
            # 한국 상품명 없으면 H열(중문/영문명) fallback
            이름 = 한국_이름 or str(row[7] or "")
            카탈로그[바코드_str] = 상품_정보(
                바코드=바코드_str,
                이름=이름,      # I열 한국명 (없으면 H열)
                브랜드=브랜드,
                제조사=제조사,
                원가=float(row[9] or 0),      # J열
                수출_usd=float(row[4] or 0),  # E열
                moq=str(row[5] or ""),        # F열
                item=str(row[2] or ""),       # C열
                model_no=str(row[3] or ""),   # D열
                번호=int(번호_val),
                원본파일=파일,
                계좌id=계좌id,
            )
        wb.close()

    @staticmethod
    def _바코드_분리(raw: str) -> tuple[str, str]:
        """
        'I열 값'에서 (바코드, 한국상품명)을 추출.
        형식: '한국상품명_바코드번호' → ('바코드번호', '한국상품명')
        형식: '바코드단독'            → ('바코드단독', '')
        """
        if "_" in raw:
            idx = raw.rfind("_")
            후보 = raw[idx + 1:].strip()
            # 후보가 숫자로만 구성된 바코드인지 확인
            if 후보 and 후보.isdigit():
                return 후보, raw[:idx].strip()
        return raw, ""

    # 파일명 suffix → 브랜드 매핑
    _전용_브랜드맵: dict[str, str] = {
        "일리온전용": "ILYON",
        "ILYON전용": "ILYON",
        "누아트전용": "NUART",
        "누아트스튜디오전용": "NUART",
        "NUART전용": "NUART",
    }

    @classmethod
    def _파일명_파싱(cls, stem: str) -> tuple[str, str]:
        """
        '富坤电器_USB 3.0 카드리더기(일리온전용)' → ('富坤电器', 'ILYON')
        '테스트_NUART 테스트(NUART전용)'          → ('테스트', 'NUART')
        '_' 없으면 ('', stem) 반환.
        괄호 suffix 없으면 '_' 뒤 첫 단어를 브랜드로 사용 (fallback).
        """
        if "_" not in stem:
            return "", stem
        제조사, 나머지 = stem.split("_", 1)

        # (xxx전용) suffix에서 브랜드 추출
        import re
        m = re.search(r'\(([^)]+전용)\)\s*$', 나머지)
        if m:
            key = m.group(1)
            if key in cls._전용_브랜드맵:
                return 제조사, cls._전용_브랜드맵[key]
            # 매핑에 없으면 괄호 안 텍스트에서 "전용" 제거 후 반환
            return 제조사, key.replace("전용", "")

        # fallback: 첫 단어
        브랜드 = 나머지.split(" ")[0]
        return 제조사, 브랜드
