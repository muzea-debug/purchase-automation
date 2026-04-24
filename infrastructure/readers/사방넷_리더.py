"""
사방넷_리더 - 사방넷단품대량수정 xlsx → 바코드별 현재고

헤더: Row 2 (Row 1 = 경로 표시, Row 3 = 설명)
데이터: Row 4~
식별: 헤더에 '사방넷상품코드' + '현재고' 포함
"""
from pathlib import Path
import openpyxl
from ._공통 import 컬럼_찾기, 안전_숫자, 안전_문자


class 사방넷_리더:

    def 읽기(self, 파일: Path) -> dict[str, float]:
        """바코드 → 현재고 dict 반환."""
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = wb.active

        헤더_행_idx = self._헤더_행_찾기(ws)
        헤더 = [ws.cell(헤더_행_idx, c).value for c in range(1, 40)]

        바코드_col = 컬럼_찾기(헤더, ["바코드"])
        현재고_col = 컬럼_찾기(헤더, ["현재고"])

        if 바코드_col is None or 현재고_col is None:
            wb.close()
            raise ValueError("사방넷 파일에서 '바코드' 또는 '현재고' 컬럼을 찾을 수 없습니다.")

        데이터_시작 = 헤더_행_idx + 2  # 헤더 + 설명행 스킵

        결과: dict[str, float] = {}
        for row in ws.iter_rows(min_row=데이터_시작, values_only=True):
            바코드 = 안전_문자(row[바코드_col] if len(row) > 바코드_col else None)
            if not 바코드 or 바코드 == '0':
                continue
            현재고 = 안전_숫자(row[현재고_col] if len(row) > 현재고_col else 0)
            # 같은 바코드 여러 행 → 합산
            결과[바코드] = 결과.get(바코드, 0.0) + 현재고

        wb.close()
        return 결과

    @staticmethod
    def _헤더_행_찾기(ws) -> int:
        """'사방넷상품코드' 또는 '바코드'가 있는 행 반환 (1-indexed)."""
        for r in range(1, 8):
            for c in range(1, 10):
                v = ws.cell(r, c).value
                if v and ('사방넷상품코드' in str(v) or ('바코드' in str(v) and '현재고' in str(ws.cell(r, c+14).value or ''))):
                    return r
        return 2  # fallback
