"""
재고현황표_리더 - 재고현황표.xlsx → 제품 마스터 dict[바코드 → 메타데이터]

헤더 구조:
  Row 1~2: 섹션 제목
  Row 3: 주요 컬럼 헤더 (바코드, 아이템, 브랜드, 상품명, 원가 …)
  Row 4: 카테고리 헤더 (시즌구분, 대분류, 소분류) + 합계 수식
  Row 5~: 데이터
"""
from pathlib import Path
import openpyxl
from ._공통 import 컬럼_찾기, 안전_숫자, 안전_문자


# 마스터 데이터에서 필요한 컬럼 헤더 키워드
_주요_컬럼 = {
    "바코드":    ["바코드"],
    "아이템":    ["아이템"],
    "브랜드":    ["브랜드"],
    "발주서명":  ["발주서명"],
    "상품명":    ["상품명"],
    "주의사항":  ["주의사항"],
    "원가_위안": ["원가", "위안"],
    "최종원가":  ["최종원가"],
    "기본_lt":   ["기본", "l/t", "lt"],
    "moq":       ["moq"],
    "안전일수":  ["안전", "일수"],
}
_카테고리_컬럼 = {
    "시즌구분": ["시즌"],
    "대분류":   ["대분류"],
    "소분류":   ["소분류"],
}


class 재고현황표_리더:

    def 읽기(self, 파일: Path) -> dict[str, dict]:
        """
        바코드 → 제품 메타데이터 dict 반환.
        카테고리(시즌구분/대분류/소분류) 포함.
        """
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = self._활성_시트(wb)

        주요_헤더_행_idx, 카테고리_헤더_행_idx = self._헤더_행_찾기(ws)

        주요_컬럼_맵 = self._컬럼_맵_생성(ws, 주요_헤더_행_idx, _주요_컬럼)
        카테고리_컬럼_맵 = self._컬럼_맵_생성(ws, 카테고리_헤더_행_idx, _카테고리_컬럼)

        바코드_col = 주요_컬럼_맵.get("바코드")
        if 바코드_col is None:
            wb.close()
            raise ValueError("재고현황표에서 '바코드' 컬럼을 찾을 수 없습니다.")

        데이터_시작_행 = max(주요_헤더_행_idx, 카테고리_헤더_행_idx) + 2  # 1-indexed, 헤더 다음 행

        결과: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=데이터_시작_행, values_only=True):
            바코드 = 안전_문자(row[바코드_col] if len(row) > 바코드_col else None)
            if not 바코드 or 바코드 == '0':
                continue

            def _get(맵, 키, 기본=''):
                col = 맵.get(키)
                if col is None or len(row) <= col:
                    return 기본
                return row[col]

            결과[바코드] = {
                "바코드":    바코드,
                "시즌구분":  안전_문자(_get(카테고리_컬럼_맵, "시즌구분")),
                "대분류":    안전_문자(_get(카테고리_컬럼_맵, "대분류")),
                "소분류":    안전_문자(_get(카테고리_컬럼_맵, "소분류")),
                "아이템":    안전_문자(_get(주요_컬럼_맵, "아이템")),
                "브랜드":    안전_문자(_get(주요_컬럼_맵, "브랜드")),
                "발주서명":  안전_문자(_get(주요_컬럼_맵, "발주서명")),
                "상품명":    안전_문자(_get(주요_컬럼_맵, "상품명")),
                "주의사항":  안전_문자(_get(주요_컬럼_맵, "주의사항")),
                "원가_위안": 안전_숫자(_get(주요_컬럼_맵, "원가_위안", 0)),
                "최종원가":  안전_숫자(_get(주요_컬럼_맵, "최종원가", 0)),
                "기본_lt":   안전_숫자(_get(주요_컬럼_맵, "기본_lt", 0)),
                "moq":       안전_숫자(_get(주요_컬럼_맵, "moq", 0)),
                "안전일수":  안전_숫자(_get(주요_컬럼_맵, "안전일수", 0)),
            }
        wb.close()
        return 결과

    # ── 내부 헬퍼 ──────────────────────────────────────────────

    @staticmethod
    def _활성_시트(wb):
        """'발주 수량 산정' 또는 '재고현황' 포함 시트 우선, 없으면 바코드 컬럼이 있는 시트."""
        우선순위 = []
        for name in wb.sheetnames:
            n = name.lower()
            if ('발주' in n and '수량' in n) or '재고현황' in n:
                우선순위.insert(0, name)
            elif name not in ('setting', 'db', '메인', 'sheet'):
                우선순위.append(name)
        # 우선순위 시트 중 '바코드' 키워드가 있는 것 선택
        for name in 우선순위:
            ws = wb[name]
            for r in range(1, 15):
                for c in range(1, 30):
                    v = ws.cell(r, c).value
                    if v and '바코드' in str(v):
                        return ws
        return wb.active

    @staticmethod
    def _헤더_행_찾기(ws) -> tuple[int, int]:
        """
        '바코드' 셀이 있는 행 → 주요 헤더 행 (1-indexed)
        '시즌구분' 셀이 있는 행 → 카테고리 헤더 행 (1-indexed)
        최대 20행까지 검색
        """
        주요 = None
        카테고리 = None
        for r in range(1, 21):
            for c in range(1, 30):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                v_str = str(v).lower().replace('\n', ' ')
                if 주요 is None and '바코드' in v_str:
                    주요 = r
                if 카테고리 is None and '시즌' in v_str:
                    카테고리 = r
            if 주요 and 카테고리:
                break
        if 주요 is None:
            주요 = 3
        if 카테고리 is None:
            카테고리 = 주요 + 1
        return 주요, 카테고리

    @staticmethod
    def _컬럼_맵_생성(ws, 헤더_행: int, 컬럼_정의: dict) -> dict[str, int]:
        """헤더 행에서 컬럼 정의에 맞는 인덱스 맵 생성 (0-indexed)."""
        헤더 = [ws.cell(헤더_행, c).value for c in range(1, ws.max_column + 1)]
        결과 = {}
        for 키, 키워드들 in 컬럼_정의.items():
            require_all = len(키워드들) > 1
            idx = 컬럼_찾기(헤더, 키워드들, require_all=require_all)
            if idx is not None:
                결과[키] = idx
        return 결과
