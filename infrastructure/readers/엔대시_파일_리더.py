"""
엔대시_파일_리더 - 아카이브 폴더의 엔대시_구글시트*.xlsx 읽기
Google Sheets 발주요청_읽기()와 동일한 (헤더, 행들) 포맷 반환
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime, date

import openpyxl


_시트_키워드 = ["발주요청", "교기관리", "交期"]


class 엔대시파일_리더:

    def 읽기(self, 파일: Path) -> tuple[list, list[list]]:
        """
        xlsx → (헤더_행, 데이터_행들)
        Google Sheets 발주요청_읽기()와 동일한 포맷.
        날짜 셀은 datetime 객체 그대로 반환 (유스케이스가 처리).
        """
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = self._시트_찾기(wb)

        헤더 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        행들: list[list] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 전체 빈 행 스킵
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            # None 값은 '' 대신 None 유지 (유스케이스의 _v() 함수와 일치)
            행들.append(list(row))

        wb.close()
        return 헤더, 행들

    def _시트_찾기(self, wb):
        for name in wb.sheetnames:
            for kw in _시트_키워드:
                if kw in name:
                    return wb[name]
        # 키워드 매칭 없으면 첫 번째 시트
        return wb.active
