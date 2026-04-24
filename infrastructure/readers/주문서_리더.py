"""
주문서_리더 - 주문서확인처리 xlsx → 바코드별 일반채널 주간판매

헤더: Row 1
식별: 헤더에 '쇼핑몰명' + '주문상태' 포함

필터:
  - 주문상태: '교환완료', '출고완료' 만 포함
  - 쇼핑몰명: '쿠팡', '로켓배송', '로켓배송(호네츠)' 제외
"""
from pathlib import Path
import openpyxl
from ._공통 import 컬럼_찾기, 안전_숫자, 안전_문자

_포함_상태 = {"교환완료", "출고완료"}
_제외_쇼핑몰 = {"쿠팡", "로켓배송", "로켓배송(호네츠)"}


class 주문서_리더:

    def 읽기(self, 파일: Path) -> dict[str, float]:
        """바코드 → 일반채널 주문수량 합계 dict 반환."""
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = wb.active

        헤더 = [ws.cell(1, c).value for c in range(1, 30)]

        바코드_col  = 컬럼_찾기(헤더, ["바코드"])
        수량_col    = 컬럼_찾기(헤더, ["수량"])
        상태_col    = 컬럼_찾기(헤더, ["주문상태"])
        쇼핑몰_col  = 컬럼_찾기(헤더, ["쇼핑몰명"])

        if 바코드_col is None or 수량_col is None:
            wb.close()
            raise ValueError("주문서 파일에서 '바코드' 또는 '수량' 컬럼을 찾을 수 없습니다.")

        결과: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 주문상태 필터
            if 상태_col is not None and len(row) > 상태_col:
                상태 = 안전_문자(row[상태_col])
                if 상태 not in _포함_상태:
                    continue

            # 쇼핑몰 필터
            if 쇼핑몰_col is not None and len(row) > 쇼핑몰_col:
                몰 = 안전_문자(row[쇼핑몰_col])
                if 몰 in _제외_쇼핑몰:
                    continue

            바코드 = 안전_문자(row[바코드_col] if len(row) > 바코드_col else None)
            if not 바코드:
                continue
            수량 = 안전_숫자(row[수량_col] if len(row) > 수량_col else 0)
            결과[바코드] = 결과.get(바코드, 0.0) + 수량

        wb.close()
        return 결과
