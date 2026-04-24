"""
forecast_리더 - A00412786-tip-*.xlsx → 바코드별 쿠팡 주간판매(최신 주)

헤더 구조:
  Row 1: 주간 시작일 (20260412, 20260419 …)  ← Col 4 부터
  Row 2: 주간 범위 (20260418~20260425 …)
  Row 3: VENDOR ID | SKU | SKU NAME | 주간예측…
  Row 4~: 데이터

식별: 파일명에 'tip' 포함
SKU→바코드 변환: nuartcompany 매핑 필요
"""
import math
from pathlib import Path
import openpyxl
from ._공통 import 안전_숫자, 안전_문자


def _사사오입(v: float) -> int:
    """Python round()는 은행가 반올림 → 사사오입(0.5 올림)으로 통일"""
    return math.floor(v + 0.5)


class Forecast_리더:

    def 읽기(self, 파일: Path, sku_바코드_맵: dict[str, str]) -> dict[str, float]:
        """바코드 → 쿠팡 주간판매(최신 주, 사사오입) dict 반환."""
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = wb.active

        # Row 1에서 최신 날짜 컬럼 찾기 (Col 4 = index 3 부터)
        최신_col = self._최신_날짜_컬럼(ws)

        # Row 3 헤더에서 SKU 컬럼 찾기
        헤더 = [ws.cell(3, c).value for c in range(1, 최신_col + 3)]
        sku_col = None
        for i, h in enumerate(헤더):
            if h and str(h).strip().upper() == 'SKU':
                sku_col = i
                break

        if sku_col is None or 최신_col is None:
            wb.close()
            raise ValueError("forecast 파일 구조를 인식할 수 없습니다.")

        결과: dict[str, float] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            sku = 안전_문자(row[sku_col] if len(row) > sku_col else None)
            if not sku:
                continue
            판매량_raw = 안전_숫자(row[최신_col] if len(row) > 최신_col else 0)
            판매량 = float(math.ceil(판매량_raw))  # 올림(round up)

            바코드 = sku_바코드_맵.get(str(sku))
            if 바코드:
                결과[바코드] = 결과.get(바코드, 0.0) + 판매량

        wb.close()
        return 결과

    @staticmethod
    def _최신_날짜_컬럼(ws) -> int | None:
        """Row 1의 Col 4부터 첫 번째 유효한 날짜 컬럼 인덱스 반환 (0-indexed).
        파일은 항상 가장 가까운 주부터 시작하므로 첫 컬럼이 정답."""
        for c in range(4, 50):
            v = ws.cell(1, c).value
            if v is None:
                continue
            try:
                int(str(v).replace('-', ''))
                return c - 1  # 0-indexed
            except (ValueError, TypeError):
                continue
        return None
