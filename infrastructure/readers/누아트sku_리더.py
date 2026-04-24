"""
누아트SKU_리더 - nuartcompany_sku_download xlsx → SKU ID ↔ 바코드 매핑

헤더: Row 1
식별: 파일명에 'nuartcompany' 포함
"""
from pathlib import Path
import openpyxl
from ._공통 import 컬럼_찾기, 안전_문자


class 누아트SKU_리더:

    def 읽기(self, 파일: Path) -> dict[str, str]:
        """SKU ID(문자열) → 바코드 dict 반환."""
        wb = openpyxl.load_workbook(파일, data_only=True)
        ws = wb.active

        헤더 = [ws.cell(1, c).value for c in range(1, 25)]
        sku_col = 컬럼_찾기(헤더, ["sku id", "sku_id"])
        바코드_col = 컬럼_찾기(헤더, ["바코드"])

        if sku_col is None or 바코드_col is None:
            wb.close()
            raise ValueError("nuartcompany 파일에서 'SKU ID' 또는 '바코드' 컬럼을 찾을 수 없습니다.")

        결과: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = 안전_문자(row[sku_col] if len(row) > sku_col else None)
            바코드 = 안전_문자(row[바코드_col] if len(row) > 바코드_col else None)
            if sku and 바코드:
                결과[sku] = 바코드

        wb.close()
        return 결과
