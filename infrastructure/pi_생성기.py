"""
PI (Proforma Invoice) xlsx 생성기

셀 위치 (새 양식 고정):
  I6  : Date (YYYY-MM-DD)
  K6  : PI No. (YYYYMMDD + po코드)
  B10 : 한국 담당자 이름
  B11 : 한국 담당자 이메일
  J8  : 중국 담당자 이름
  J9  : 중국 담당자 이메일
  J10 : 중국 담당자 연락처
  14~ : 제품 행 (A~K)
  20  : TOTAL 행
  21~ : 디포짓 행 (K열 = 금액)
  25  : Balance 행
  B28 : Shipment
  B29 : Payment
  B30~: Bank information (각 항목 별도 행)
  A45:D48 : 누아트 도장
  I45:K48 : 제조사 도장
"""
from __future__ import annotations
import shutil
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font

import config

_YELLOW   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_RED_BOLD = Font(color="FF0000", bold=True)
_NO_FILL  = PatternFill(fill_type=None)
_NORMAL   = Font()

# 제품 헤더 행 / TOTAL 행 고정 위치
_HEADER_ROW  = 13
_TOTAL_ROW   = 20
# 디포짓 행 고정 (최대 4차)
_DEP_ROWS    = [21, 22, 23, 24]
_BALANCE_ROW = 25
# 하단 고정 행
_SHIPMENT_ROW   = 28
_PAYMENT_ROW    = 29
_BANK_START_ROW = 30
# 도장 고정 앵커
_BUYER_STAMP_ANCHOR  = "A45"
_SELLER_STAMP_ANCHOR = "I45"


class PI생성기:

    def __init__(self):
        self._템플릿 = config.PI_템플릿
        config.PI_DIR.mkdir(parents=True, exist_ok=True)

    def 생성(
        self,
        po코드: str,
        제조사_계좌: dict,
        제품목록: list[dict],
        디포짓_목록: list[dict],
        담당자: str,
        중국담당자: str = "",
        담당자이메일: str = "",
        중국이메일: str = "",
        중국연락처: str = "",
        max_lt: int = 0,
        pi_번호: Optional[str] = None,
    ) -> Path:
        if not self._템플릿.exists():
            raise FileNotFoundError(f"PI 템플릿 없음: {self._템플릿}")

        오늘 = date.today()
        if pi_번호 is None:
            pi_번호 = f"{오늘.strftime('%Y%m%d')}{po코드}"

        회사명_short = (
            제조사_계좌.get("회사명_영문", "")[:20]
            .replace("/", "-").replace("\\", "-")
        )
        출력_파일명 = f"{오늘.strftime('%y%m%d')}_{po코드}_{회사명_short}.xlsx"
        출력_경로 = config.PI_DIR / po코드
        출력_경로.mkdir(parents=True, exist_ok=True)
        출력_파일 = 출력_경로 / 출력_파일명

        shutil.copy2(self._템플릿, 출력_파일)
        wb = openpyxl.load_workbook(출력_파일)
        ws = wb.active

        총액 = sum(
            int(p.get("qty", 0)) * float(p.get("unit_price", 0))
            for p in 제품목록
        )
        완료_디포짓_합 = sum(
            float(d.get("amount") or (총액 * float(d.get("pct", 0))))
            for d in 디포짓_목록
            if d.get("완료")
        )

        self._fill_header(ws, 오늘, pi_번호, 담당자, 담당자이메일,
                          중국담당자, 중국이메일, 중국연락처, 제조사_계좌)
        self._fill_products(ws, 제품목록, 총액, 완료_디포짓_합)
        self._fill_deposits(ws, 디포짓_목록, 총액, 완료_디포짓_합)
        self._fill_shipment(ws, 오늘, max_lt)
        self._fill_payment(ws, 디포짓_목록)
        self._fill_bank_info(ws, 제조사_계좌)
        self._insert_stamps(ws, 제조사_계좌)

        wb.save(출력_파일)
        return 출력_파일

    # ── 셀 쓰기 헬퍼 ──────────────────────────────────────────────

    def _set(self, ws, row: int, col: int, value):
        try:
            ws.cell(row, col).value = value
        except AttributeError:
            pass

    # ── 헤더 ──────────────────────────────────────────────────────

    def _fill_header(self, ws, 오늘: date, pi_번호: str,
                     담당자: str, 담당자이메일: str,
                     중국담당자: str, 중국이메일: str, 중국연락처: str,
                     계좌: dict):
        # I6: 날짜, K6: PI 번호
        self._set(ws, 6, 9,  오늘.strftime("%Y-%m-%d"))
        self._set(ws, 6, 11, pi_번호)

        # 한국 담당자
        self._set(ws, 10, 2, 담당자)
        self._set(ws, 11, 2, 담당자이메일)

        # 중국 담당자 (파라미터 우선, 없으면 계좌 DB)
        cn_name  = 중국담당자  or 계좌.get("seller_name", "")
        cn_email = 중국이메일  or 계좌.get("seller_email", "")
        cn_phone = 중국연락처  or 계좌.get("seller_phone", "")
        self._set(ws, 8,  10, cn_name)
        self._set(ws, 9,  10, cn_email)
        self._set(ws, 10, 10, cn_phone)

        # J11 (Office Add) 클리어
        self._set(ws, 11, 10, None)
        self._set(ws, 11, 11, None)

    # ── 제품 행 ────────────────────────────────────────────────────

    def _fill_products(self, ws, 제품목록: list[dict], 총액: float, 완료_디포짓_합: float):
        # 제품 행이 부족하면 TOTAL 행 위에 삽입
        n_needed = len(제품목록)
        n_exist  = _TOTAL_ROW - _HEADER_ROW - 1   # 기본 템플릿 제품 행 수

        total_row = _TOTAL_ROW
        if n_needed > n_exist:
            ws.insert_rows(total_row, n_needed - n_exist)
            total_row += (n_needed - n_exist)

        for i, prod in enumerate(제품목록):
            r   = _HEADER_ROW + 1 + i
            qty  = int(prod.get("qty", 0))
            unit = float(prod.get("unit_price", 0))
            amt  = round(qty * unit, 2)

            # Already Paid: 완료 디포짓을 금액 비례로 배분
            paid = round(완료_디포짓_합 * (amt / 총액), 2) if 총액 > 0 else 0.0
            bal  = round(amt - paid, 2)

            self._set(ws, r, 1,  i + 1)
            self._set(ws, r, 2,  prod.get("po_no", ""))
            self._set(ws, r, 3,  prod.get("brand", ""))
            self._set(ws, r, 4,  prod.get("description_en", "") or prod.get("description", ""))
            self._set(ws, r, 5,  prod.get("이름", "") or prod.get("product_name_kr", ""))
            self._set(ws, r, 6,  str(prod.get("barcode", "")))
            self._set(ws, r, 7,  qty)
            self._set(ws, r, 8,  unit)
            self._set(ws, r, 9,  amt)
            self._set(ws, r, 10, paid)
            self._set(ws, r, 11, bal)

        # 남은 행 클리어
        for r in range(_HEADER_ROW + 1 + n_needed, total_row):
            for c in range(1, 12):
                self._set(ws, r, c, None)

        # TOTAL 행
        total_qty  = sum(int(p.get("qty", 0)) for p in 제품목록)
        total_bal  = round(총액 - 완료_디포짓_합, 2)
        self._set(ws, total_row, 7,  total_qty)
        self._set(ws, total_row, 9,  round(총액, 2))
        self._set(ws, total_row, 10, round(완료_디포짓_합, 2))
        self._set(ws, total_row, 11, max(total_bal, 0))

    # ── 디포짓 ─────────────────────────────────────────────────────

    def _fill_deposits(self, ws, 디포짓_목록: list[dict], 총액: float, 완료_디포짓_합: float):
        n = len(디포짓_목록)

        # 사용하는 차수 채우기
        for i, row in enumerate(_DEP_ROWS):
            if i >= n:
                break
            dep = 디포짓_목록[i]
            pct = float(dep.get("pct", 0))
            amt = float(dep.get("amount") or round(총액 * pct, 2))
            self._set(ws, row, 11, round(amt, 2))

            if dep.get("_현재"):
                for c in range(9, 12):
                    ws.cell(row, c).fill = _YELLOW
                    ws.cell(row, c).font = _RED_BOLD
            else:
                for c in range(9, 12):
                    ws.cell(row, c).fill = _NO_FILL
                    ws.cell(row, c).font = _NORMAL

        # 미사용 차수 행 삭제 (아래서 위로)
        for i in sorted(range(n, len(_DEP_ROWS)), reverse=True):
            ws.delete_rows(_DEP_ROWS[i])

        # Balance 행: 삭제 후 실제 행 번호 재계산
        balance_row = _BALANCE_ROW - (len(_DEP_ROWS) - n)
        잔액 = max(round(총액 - 완료_디포짓_합, 2), 0)
        self._set(ws, balance_row, 10, round(완료_디포짓_합, 2))
        self._set(ws, balance_row, 11, 잔액)

    # ── Shipment ────────────────────────────────────────────────────

    def _fill_shipment(self, ws, 오늘: date, max_lt: int):
        if max_lt > 0:
            예정일 = 오늘 + timedelta(days=max_lt)
            # 주말 건너뜀
            while 예정일.weekday() >= 5:
                예정일 += timedelta(days=1)
            text = (f"About {예정일.strftime('%Y-%m-%d')}, "
                    f"within {max_lt} working days after confirmation of payment.")
        else:
            text = "Within agreed working days after confirmation of payment."
        self._set(ws, _SHIPMENT_ROW, 2, text)

    # ── Payment ─────────────────────────────────────────────────────

    def _fill_payment(self, ws, 디포짓_목록: list[dict]):
        ORDINALS = ["1ST", "2ND", "3RD", "4TH"]
        n = len(디포짓_목록)
        parts = []
        for i, dep in enumerate(디포짓_목록):
            pct = float(dep.get("pct", 0))
            pct_int = round(pct * 100)
            ord_label = ORDINALS[i] if i < len(ORDINALS) else f"{i+1}TH"
            parts.append(f"{pct_int}% {ord_label} Deposit")
        if parts:
            text = ", ".join(parts) + ". Shipping schedule to be discussed separately."
        else:
            text = "Payment terms to be discussed separately."
        self._set(ws, _PAYMENT_ROW, 2, text)

    # ── Bank Information ────────────────────────────────────────────

    def _fill_bank_info(self, ws, 계좌: dict):
        항목들 = [
            ("Beneficiary Name",    계좌.get("beneficiary_name") or 계좌.get("회사명_영문", "")),
            ("Beneficiary Address", 계좌.get("beneficiary_address", "")),
            ("Country / City",      f"{계좌.get('country','')} / {계좌.get('city','')}".strip(" /")),
            ("Account Number",      계좌.get("account_number", "")),
            ("SWIFT / BIC",         계좌.get("swift", "")),
            ("Bank Name",           계좌.get("bank_name", "")),
            ("Bank Address",        계좌.get("bank_address", "")),
            ("Sort Code",           계좌.get("sort_code", "")),
            ("Branch Code",         계좌.get("branch_code", "")),
        ]

        r = _BANK_START_ROW
        for label, value in 항목들:
            if not value or value.strip(" /") == "":
                continue
            self._set(ws, r, 1, f"{label} :")
            self._set(ws, r, 2, value)
            r += 1

        # 남은 행 클리어
        for clear_r in range(r, _BANK_START_ROW + 12):
            self._set(ws, clear_r, 1, None)
            self._set(ws, clear_r, 2, None)

    # ── 도장 ───────────────────────────────────────────────────────

    def _insert_stamps(self, ws, 계좌: dict):
        ws._images.clear()

        # 누아트 도장 → A45
        if config.누아트_도장.exists():
            try:
                img = XLImage(str(config.누아트_도장))
                img.width, img.height = 130, 85
                ws.add_image(img, _BUYER_STAMP_ANCHOR)
            except Exception as e:
                print(f"[PI생성기] 누아트 도장 삽입 실패: {e}")

        # 제조사 도장 → I45
        도장_파일 = 계좌.get("도장_파일", "")
        if not 도장_파일:
            return
        도장_경로 = config.업체도장_DIR / Path(도장_파일).name
        if not 도장_경로.exists():
            도장_경로 = config.BASE_DIR / 도장_파일
        if not 도장_경로.exists():
            return
        try:
            img = XLImage(str(도장_경로))
            img.width, img.height = 130, 85
            ws.add_image(img, _SELLER_STAMP_ANCHOR)
        except Exception as e:
            print(f"[PI생성기] 제조사 도장 삽입 실패: {e}")
