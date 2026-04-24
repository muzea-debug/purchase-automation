"""
엑셀 시트 저장소 - Google Sheets API 없이 로컬 Excel 파일로 동일 기능 제공
구글 시트 대체용/ 폴더의 엑셀 파일 사용

탭 구조:
  엔대시                         ← 발주 기록 로그 (발주_기록 write)
  중국 발주요청리스트(交期管理表) ← 재고현황 소스 (발주요청_읽기 read)
  재고현황_아카이브               ← 아카이브
  제품DB                         ← 제품 마스터 (바코드·브랜드·카테고리 등)
"""
import re
from pathlib import Path
from datetime import date

import openpyxl

_탭_엔대시   = "엔대시"
_탭_제품DB   = "제품DB"


class 엑셀시트저장소:

    def __init__(self, 파일_경로: Path):
        self._파일 = Path(파일_경로)

    # ── 내부 ────────────────────────────────────────────────────

    def _읽기_wb(self):
        return openpyxl.load_workbook(self._파일, data_only=True)

    def _쓰기_wb(self):
        return openpyxl.load_workbook(self._파일)

    # ── GoogleSheets저장소 호환 인터페이스 ───────────────────────

    def 마지막_po번호_조회(self, prefix: str) -> int:
        try:
            wb = self._읽기_wb()
            if _탭_엔대시 not in wb.sheetnames:
                wb.close(); return 0
            ws = wb[_탭_엔대시]
            pat = re.compile(rf'^{re.escape(prefix)}\((\d+)\)$')
            max_num = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 2 and row[2]:
                    m = pat.match(str(row[2]).strip())
                    if m:
                        max_num = max(max_num, int(m.group(1)))
            wb.close()
            return max_num
        except Exception as e:
            print(f"[엑셀저장소] po번호 조회 실패: {e}")
            return 0

    def 발주_기록(self, 바코드: str, po코드: str, 수량: int,
                  제품명: str, 제조사: str) -> None:
        try:
            wb = self._쓰기_wb()
            if _탭_엔대시 not in wb.sheetnames:
                wb.create_sheet(_탭_엔대시)
            ws = wb[_탭_엔대시]
            오늘 = date.today()
            발주일자 = f"{오늘.year}. {오늘.month}. {오늘.day}"
            ws.append([발주일자, 바코드, po코드, 수량, "", 제품명, "", "", "", "", 제조사])
            wb.save(self._파일)
            wb.close()
        except Exception as e:
            print(f"[엑셀저장소] 발주 기록 실패: {e}")

    def 발주요청_읽기(self, 시트명: str) -> tuple[list, list]:
        try:
            wb = self._읽기_wb()
            if 시트명 not in wb.sheetnames:
                wb.close(); return [], []
            ws = wb[시트명]
            모든행 = list(ws.iter_rows(values_only=True))
            wb.close()
            if not 모든행:
                return [], []
            헤더 = [str(v) if v is not None else "" for v in 모든행[0]]
            데이터 = [list(r) for r in 모든행[1:] if any(v is not None for v in r)]
            return 헤더, 데이터
        except Exception as e:
            print(f"[엑셀저장소] 발주요청 읽기 실패: {e}")
            return [], []

    def 아카이브_추가(self, 아카이브_시트명: str, 행들: list[list]) -> None:
        if not 행들:
            return
        try:
            wb = self._쓰기_wb()
            if 아카이브_시트명 not in wb.sheetnames:
                wb.create_sheet(아카이브_시트명)
            ws = wb[아카이브_시트명]
            for row in 행들:
                ws.append(row)
            wb.save(self._파일)
            wb.close()
        except Exception as e:
            print(f"[엑셀저장소] 아카이브 추가 실패: {e}")

    # ── 추가 기능 ────────────────────────────────────────────────

    def 제품DB_수정(self, 바코드: str, 변경_필드: dict) -> bool:
        """제품DB 탭에서 바코드 행의 특정 필드를 업데이트."""
        _필드_키워드 = {
            "moq":       ["moq"],
            "기본_lt":   ["기본lt", "기본_lt", "l/t", "lt"],
            "원가_위안": ["제품원가", "원가위안", "원가_위안"],
            "최종원가":  ["최종원가"],
            "발주서명":  ["발주서명"],
            "안전일수":  ["안전일수"],
            "주의사항":  ["주의사항"],
            "브랜드":    ["브랜드"],
        }
        try:
            wb = self._쓰기_wb()
            if _탭_제품DB not in wb.sheetnames:
                wb.close(); return False
            ws = wb[_탭_제품DB]
            rows = list(ws.iter_rows())
            if not rows:
                wb.close(); return False

            헤더 = [str(c.value).strip().lower().replace(" ", "").replace("\n", "") if c.value else "" for c in rows[0]]

            def _find_col(*keywords):
                for kw in keywords:
                    kw_n = kw.lower().replace(" ", "")
                    for i, h in enumerate(헤더):
                        if kw_n in h:
                            return i
                return None

            바코드_c = _find_col("바코드")
            if 바코드_c is None:
                wb.close(); return False

            col_map = {field: _find_col(*kws) for field, kws in _필드_키워드.items()}

            갱신됨 = False
            for row in rows[1:]:
                cell_val = str(row[바코드_c].value or "").strip()
                if cell_val == 바코드:
                    for field, val in 변경_필드.items():
                        c = col_map.get(field)
                        if c is not None:
                            row[c].value = val
                    갱신됨 = True
                    break

            if 갱신됨:
                wb.save(self._파일)
            wb.close()
            return 갱신됨
        except Exception as e:
            print(f"[엑셀저장소] 제품DB 수정 실패: {e}")
            return False

    def 제품DB_읽기(self) -> dict[str, dict]:
        """제품DB 탭 → 바코드별 제품 메타데이터 dict."""
        try:
            wb = self._읽기_wb()
            if _탭_제품DB not in wb.sheetnames:
                wb.close(); return {}
            ws = wb[_탭_제품DB]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(rows) < 2:
                return {}

            헤더 = [str(v).strip().lower().replace(" ", "").replace("\n", "") if v else "" for v in rows[0]]

            def _col(*keywords):
                for i, h in enumerate(헤더):
                    if any(k.lower().replace(" ", "") in h for k in keywords):
                        return i
                return None

            바코드_c   = _col("바코드")
            브랜드_c   = _col("브랜드")
            발주서명_c = _col("발주서명")
            상품명_c   = _col("상품명")
            moq_c      = _col("moq")
            원가위안_c = _col("제품원가", "원가위안", "원가_위안")
            최종원가_c = _col("최종원가")
            시즌_c     = _col("시즌구분", "시즌")
            대분류_c   = _col("대분류")
            소분류_c   = _col("소분류")
            아이템_c   = _col("아이템")
            안전일수_c = _col("안전일수")
            lt_c       = _col("기본lt", "기본_lt", "l/t", "lt")
            주의사항_c = _col("주의사항")

            if 바코드_c is None:
                return {}

            def _v(row, col, default=""):
                if col is None or col >= len(row):
                    return default
                return row[col] if row[col] is not None else default

            def _n(row, col):
                v = _v(row, col, 0)
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return 0.0

            def _moq_max(row, col):
                """'15~20', '35-45', '100' → 최대값 숫자 반환."""
                import re as _re
                v = str(_v(row, col, "0"))
                nums = _re.findall(r'\d+', v)
                if not nums:
                    return 0.0
                return float(max(int(n) for n in nums))

            결과 = {}
            for row in rows[1:]:
                바코드 = str(_v(row, 바코드_c, "")).strip()
                if not 바코드:
                    continue
                결과[바코드] = {
                    "바코드":    바코드,
                    "브랜드":    str(_v(row, 브랜드_c)),
                    "발주서명":  str(_v(row, 발주서명_c)),
                    "상품명":    str(_v(row, 상품명_c)),
                    "moq":           str(_v(row, moq_c)),
                    "발주확정용_moq": _moq_max(row, moq_c),
                    "원가_위안": _n(row, 원가위안_c),
                    "최종원가":  _n(row, 최종원가_c),
                    "시즌구분":  str(_v(row, 시즌_c)),
                    "대분류":    str(_v(row, 대분류_c)),
                    "소분류":    str(_v(row, 소분류_c)),
                    "아이템":    str(_v(row, 아이템_c)),
                    "안전일수":  _n(row, 안전일수_c),
                    "기본_lt":   _n(row, lt_c),
                    "주의사항":  str(_v(row, 주의사항_c)),
                }
            return 결과
        except Exception as e:
            print(f"[엑셀저장소] 제품DB 읽기 실패: {e}")
            return {}
