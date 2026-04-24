"""
로컬 Excel 시트 초기화 - 구글 시트 대체용 엑셀 파일 생성

탭 생성 순서:
  1. 엔대시                         ← 발주 기록 (헤더만)
  2. 중국 발주요청리스트(交期管理表) ← 엔대시 소스 파일에서 복사
  3. 재고현황_아카이브               ← 헤더만
  4. 제품DB                         ← 주문프로그램 DB + 재고현황표 카테고리 병합
"""
from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

_헤더_색  = "1A1A2E"
_헤더_폰트 = Font(bold=True, color="FFFFFF", size=10)
_헤더_채움 = PatternFill("solid", fgColor=_헤더_색)
_가운데 = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _헤더_스타일(ws, 행번호=1):
    for cell in ws[행번호]:
        cell.font = _헤더_폰트
        cell.fill = _헤더_채움
        cell.alignment = _가운데


# ── 1. 엔대시 탭 (발주 기록) ────────────────────────────────────

_엔대시_헤더 = [
    "발주일자", "바코드", "발주번호(PO)", "쿠팡발주수량", "일반수량",
    "제품명", "비고", "중국출발일자", "누아트입고일자", "14일경과", "제조사",
]

def _엔대시_탭(wb):
    ws = wb.create_sheet("엔대시")
    ws.append(_엔대시_헤더)
    _헤더_스타일(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["F"].width = 30
    return ws


# ── 2. 중국 발주요청리스트 탭 ────────────────────────────────────

def _발주요청_탭(wb, 소스_파일: Path | None):
    시트명 = "중국 발주요청리스트(交期管理表)"
    ws = wb.create_sheet(시트명)

    if 소스_파일 and 소스_파일.exists():
        try:
            src = openpyxl.load_workbook(소스_파일, data_only=True)
            if 시트명 in src.sheetnames:
                src_ws = src[시트명]
                for row in src_ws.iter_rows(values_only=True):
                    ws.append(list(row))
                _헤더_스타일(ws)
                print(f"  [발주요청] {src_ws.max_row - 1}행 복사 완료")
            else:
                # 첫 번째 시트 사용
                src_ws = src.active
                for row in src_ws.iter_rows(values_only=True):
                    ws.append(list(row))
                _헤더_스타일(ws)
                print(f"  [발주요청] active 시트에서 {src_ws.max_row - 1}행 복사")
            src.close()
        except Exception as e:
            print(f"  [발주요청] 소스 읽기 실패 ({e}), 빈 헤더로 생성")
            _빈_발주요청_헤더(ws)
    else:
        _빈_발주요청_헤더(ws)
    return ws


def _빈_발주요청_헤더(ws):
    ws.append([
        "쿠팡채널 중국 재고", "일반채널 중국 재고", "발주일자", "바코드",
        "발주번호", "쿠팡 발주수량", "일반건 수량", "제품명",
        "비고/입고예정일", "중국 출발일자", "누아트 입고일자",
        "발주 14일 경과 품목", "제조사", "출고여부",
    ])
    _헤더_스타일(ws)


# ── 3. 아카이브 탭 ─────────────────────────────────────────────

def _아카이브_탭(wb):
    ws = wb.create_sheet("재고현황_아카이브")
    ws.append(["날짜", "바코드", "제품명", "재고상태", "재고합계", "발주량산출"])
    _헤더_스타일(ws)
    return ws


# ── 4. 제품DB 탭 ───────────────────────────────────────────────

_제품DB_헤더 = [
    "바코드", "브랜드", "발주서명", "상품명", "MOQ",
    "제품원가(위안)", "최종원가",
    "시즌구분", "대분류", "소분류", "아이템",
    "안전일수", "기본_lt", "주의사항",
]

def _제품DB_탭(wb, 주문프로그램_파일: Path | None, 재고현황표_파일: Path | None):
    ws = wb.create_sheet("제품DB")
    ws.append(_제품DB_헤더)
    _헤더_스타일(ws)

    # 카테고리 맵 (바코드 → 카테고리 정보) — 재고현황표에서
    카테고리_맵 = _재고현황표_카테고리(재고현황표_파일)

    # 주문프로그램 DB에서 제품 목록
    제품_목록 = _주문프로그램_DB(주문프로그램_파일)

    count = 0
    for 바코드, p in 제품_목록.items():
        cat = 카테고리_맵.get(바코드, {})
        ws.append([
            바코드,
            p.get("브랜드", ""),
            p.get("발주서명", ""),
            p.get("상품명", ""),
            p.get("moq", ""),
            p.get("원가", ""),
            p.get("최종원가", ""),
            cat.get("시즌구분", ""),
            cat.get("대분류", ""),
            cat.get("소분류", ""),
            cat.get("아이템", ""),
            cat.get("안전일수", ""),
            cat.get("기본_lt", ""),
            cat.get("주의사항", ""),
        ])
        count += 1

    # 컬럼 너비
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A2"
    print(f"  [제품DB] {count}개 제품 기록")
    return ws


def _주문프로그램_DB(파일: Path | None) -> dict[str, dict]:
    if not 파일 or not 파일.exists():
        print("  [제품DB] 주문프로그램 파일 없음, 빈 DB로 생성")
        return {}
    try:
        wb = openpyxl.load_workbook(파일, data_only=True, keep_vba=True)
        ws = wb["DB"] if "DB" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {}

        헤더 = [str(v).strip().lower() if v else "" for v in rows[0]]

        def _c(*keys):
            for i, h in enumerate(헤더):
                if any(k.lower() in h for k in keys):
                    return i
            return None

        바코드_c   = _c("바코드")
        브랜드_c   = _c("브랜드")
        발주서명_c = _c("발주서명")
        상품명_c   = _c("상품명")
        moq_c      = _c("moq", "최소구매")
        원가_c     = _c("제품원가", "원가")
        최종원가_c = _c("최종원가")

        if 바코드_c is None:
            return {}

        결과 = {}
        for row in rows[1:]:
            def _v(col, d=""):
                if col is None or col >= len(row):
                    return d
                return row[col] if row[col] is not None else d

            바코드 = str(_v(바코드_c, "")).strip()
            if not 바코드 or 바코드 in 결과:
                continue
            결과[바코드] = {
                "브랜드":   str(_v(브랜드_c)),
                "발주서명": str(_v(발주서명_c)),
                "상품명":   str(_v(상품명_c)),
                "moq":      _v(moq_c, ""),
                "원가":     _v(원가_c, ""),
                "최종원가": _v(최종원가_c, ""),
            }
        print(f"  [제품DB] 주문프로그램 DB에서 {len(결과)}개 제품 로드")
        return 결과
    except Exception as e:
        print(f"  [제품DB] 주문프로그램 읽기 실패: {e}")
        return {}


def _재고현황표_카테고리(파일: Path | None) -> dict[str, dict]:
    """재고현황표에서 바코드별 카테고리 정보 추출."""
    if not 파일 or not 파일.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(파일, data_only=True)
        # 발주 수량 산정 시트 찾기
        ws = None
        for name in wb.sheetnames:
            if "발주" in name and "수량" in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # 헤더 행 찾기 (바코드 있는 행)
        바코드_row = None
        시즌_row   = None
        for r in range(1, 15):
            for c in range(1, 30):
                v = ws.cell(r, c).value
                if v and "바코드" in str(v):
                    바코드_row = r
                if v and "시즌" in str(v):
                    시즌_row = r
            if 바코드_row and 시즌_row:
                break

        if 바코드_row is None:
            wb.close()
            return {}

        if 시즌_row is None:
            시즌_row = 바코드_row + 1

        # 컬럼 인덱스
        def _find_col(row_idx, *keys):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row_idx, c).value
                if v and any(k.lower() in str(v).lower() for k in keys):
                    return c - 1  # 0-indexed
            return None

        바코드_c  = _find_col(바코드_row, "바코드")
        시즌_c    = _find_col(시즌_row, "시즌")
        대분류_c  = _find_col(시즌_row, "대분류")
        소분류_c  = _find_col(시즌_row, "소분류")
        아이템_c  = _find_col(바코드_row, "아이템")
        안전_c    = _find_col(바코드_row, "안전", "일수")
        lt_c      = _find_col(바코드_row, "l/t", "lt", "기본")
        주의사항_c = _find_col(바코드_row, "주의사항")

        데이터_시작 = max(바코드_row, 시즌_row) + 2

        def _v(row, col, d=""):
            if col is None or col >= len(row):
                return d
            return row[col] if row[col] is not None else d

        결과 = {}
        for row in ws.iter_rows(min_row=데이터_시작, values_only=True):
            if 바코드_c is None:
                break
            bc = str(_v(row, 바코드_c, "")).strip()
            if not bc:
                continue
            결과[bc] = {
                "시즌구분":  str(_v(row, 시즌_c)),
                "대분류":    str(_v(row, 대분류_c)),
                "소분류":    str(_v(row, 소분류_c)),
                "아이템":    str(_v(row, 아이템_c)),
                "안전일수":  _v(row, 안전_c, ""),
                "기본_lt":   _v(row, lt_c, ""),
                "주의사항":  str(_v(row, 주의사항_c)),
            }
        wb.close()
        print(f"  [제품DB] 재고현황표에서 {len(결과)}개 카테고리 로드")
        return 결과
    except Exception as e:
        print(f"  [제품DB] 재고현황표 카테고리 읽기 실패: {e}")
        return {}


# ── 공개 API ───────────────────────────────────────────────────

def 로컬시트_초기화(
    저장_경로: Path,
    엔대시_소스: Path | None = None,
    주문프로그램: Path | None = None,
    재고현황표: Path | None = None,
    덮어쓰기: bool = False,
) -> bool:
    """
    로컬 Excel 시트 초기화.
    덮어쓰기=False면 파일이 이미 있을 때 건너뜀.
    Returns: True=생성됨, False=건너뜀
    """
    if 저장_경로.exists() and not 덮어쓰기:
        print(f"[로컬시트] 이미 존재: {저장_경로}")
        return False

    print(f"[로컬시트] 초기화 시작: {저장_경로}")
    저장_경로.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _엔대시_탭(wb)
    _발주요청_탭(wb, 엔대시_소스)
    _아카이브_탭(wb)
    _제품DB_탭(wb, 주문프로그램, 재고현황표)

    wb.save(저장_경로)
    wb.close()
    print(f"[로컬시트] 완료: {저장_경로}")
    return True
