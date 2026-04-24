"""
xlsx 저장소 - 발주서 엔티티 → xlsx BytesIO 반환

행 구조:
  1 - 빈 행
  2 - 제목
  3 - 빈 행
  4 - 한국 요청 도착일 / 중국지사 도착일
  5 - 패키지원가 / 부자재원가 / 리드타임
  6 - 헤더
  7+ - 제품 행
  하단(≥15행) - 포장 요구사항
"""
from pathlib import Path
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from domain.entities.발주서 import 발주서


# ── 스타일 상수 ──────────────────────────────────────────────────

_채우기_도착일    = PatternFill("solid", fgColor="B4C7E7")
_채우기_원가레이블 = PatternFill("solid", fgColor="EDEDED")
_채우기_헤더_좌   = PatternFill("solid", fgColor="EDEDED")
_채우기_헤더_우   = PatternFill("solid", fgColor="FFF2CC")
_채우기_포장레이블 = PatternFill("solid", fgColor="D6E4F7")

_제목_폰트   = Font(name="맑은 고딕", bold=True, size=11)
_헤더_폰트   = Font(name="맑은 고딕", bold=True, size=9)
_레이블_폰트  = Font(name="맑은 고딕", size=9)
_본문_폰트   = Font(name="맑은 고딕", size=9)
_중앙정렬    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_좌측정렬    = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _테두리():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _col_px(ws, col_idx: int) -> float:
    """열 너비를 픽셀로 변환 (근사: 1 char ≈ 7px + 5px 패딩)"""
    letter = get_column_letter(col_idx)
    ch = ws.column_dimensions[letter].width or 8
    return ch * 7 + 5


def _row_px(ws, row_idx: int) -> float:
    """행 높이를 픽셀로 변환 (pt → px at 96dpi)"""
    pt = ws.row_dimensions[row_idx].height or 15
    return pt * 96 / 72


def _이미지_중앙삽입(ws, path: str, col_idx: int, row_idx: int,
                    img_w: int, img_h: int, span_cols: int = 1,
                    span_rows: int = 1, fit: bool = False):
    """
    이미지를 셀 범위 중앙에 삽입.
    fit=True 이면 셀 크기에 맞추되 원본 비율 유지.
    """
    try:
        cell_w = sum(_col_px(ws, col_idx + k) for k in range(span_cols))
        cell_h = sum(_row_px(ws, row_idx + k) for k in range(span_rows))

        if fit:
            # 원본 이미지 비율로 셀에 맞춤 (여백 8px)
            pad = 8
            max_w = cell_w - pad
            max_h = cell_h - pad
            ratio = min(max_w / img_w, max_h / img_h)
            img_w = int(img_w * ratio)
            img_h = int(img_h * ratio)

        col_off = max(0.0, (cell_w - img_w) / 2)
        row_off = max(0.0, (cell_h - img_h) / 2)

        img = XLImage(path)
        img.width, img.height = img_w, img_h

        marker = AnchorMarker(
            col=col_idx - 1,
            colOff=int(pixels_to_EMU(col_off)),
            row=row_idx - 1,
            rowOff=int(pixels_to_EMU(row_off)),
        )
        size = XDRPositiveSize2D(
            int(pixels_to_EMU(img_w)),
            int(pixels_to_EMU(img_h)),
        )
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception:
        pass


class Xlsx저장소:

    def 저장_메모리(self, 발주: 발주서) -> tuple[io.BytesIO, str]:
        """워크북을 메모리에 생성하여 (BytesIO, 파일명) 반환"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{발주.브랜드} {발주.제품명}"

        self._열_너비_설정(ws)
        self._행2_제목(ws, 발주)
        self._행4_도착일(ws, 발주)
        self._행5_원가_리드타임(ws, 발주)
        self._행6_헤더(ws)
        if 발주.브랜드 in ("ILYON", "일리온", "누아트스튜디오"):
            self._외박스_섹션(ws, 발주)
        마지막행 = self._제품행_작성(ws, 발주)
        포장_시작행 = max(마지막행 + 2, 15)
        self._포장이미지_삽입(ws, 발주, 포장_시작행)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, 발주.파일명

    # ── 열 너비 ──────────────────────────────────────────────────

    def _열_너비_설정(self, ws):
        widths = {
            "A": 5.1,  "B": 6.9,  "C": 12.6, "D": 20.0,
            "E": 12.6, "F": 12.6, "G": 25.8, "H": 32.8,
            "I": 55.8, "J": 10.6, "K": 13.3, "L": 14.4,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # ── 행2: 제목 ────────────────────────────────────────────────

    def _행2_제목(self, ws, 발주: 발주서):
        c = ws["A2"]
        c.value = f"@{발주.브랜드} {발주.제품명} 발주서"
        c.font = _제목_폰트
        c.alignment = _좌측정렬
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 8

    # ── 행4: 도착일 ──────────────────────────────────────────────

    def _행4_도착일(self, ws, 발주: 발주서):
        ws.merge_cells("B4:C4")
        ws.merge_cells("E4:F4")

        ws["B4"] = "한국 요청 도착일\n韩国到达"
        ws["B4"].font = _레이블_폰트
        ws["B4"].fill = _채우기_도착일
        ws["B4"].alignment = _중앙정렬
        ws["B4"].border = _테두리()

        ws["D4"].border = _테두리()

        ws["E4"] = "중국지사 도착일\n到达中国支社日"
        ws["E4"].font = _레이블_폰트
        ws["E4"].fill = _채우기_도착일
        ws["E4"].alignment = _중앙정렬
        ws["E4"].border = _테두리()

        ws["G4"] = 발주.한국_요청_도착일 or ""
        ws["G4"].font = _본문_폰트
        ws["G4"].alignment = _중앙정렬
        ws["G4"].border = _테두리()

        for addr in ("C4", "F4"):
            ws[addr].border = _테두리()

        ws.row_dimensions[4].height = 35

    # ── 행5: 원가 / 리드타임 ─────────────────────────────────────

    def _행5_원가_리드타임(self, ws, 발주: 발주서):
        pairs = [
            ("B5", "패키지 원가 (CNY)", _채우기_원가레이블),
            ("C5", 발주.패키지원가,     None),
            ("D5", "부자재 원가 (CNY)", _채우기_원가레이블),
            ("E5", 발주.부자재원가,     None),
            ("F5", "리드타임",          _채우기_원가레이블),
            ("G5", 발주.리드타임,       None),
        ]
        for addr, val, fill in pairs:
            c = ws[addr]
            c.value = val
            c.font = _레이블_폰트 if fill else _본문_폰트
            c.alignment = _중앙정렬
            c.border = _테두리()
            if fill:
                c.fill = fill
        if 발주.계좌id:
            ws["J5"] = 발주.계좌id
        ws.row_dimensions[5].height = 45

    # ── 행6: 헤더 ────────────────────────────────────────────────

    def _행6_헤더(self, ws):
        좌측_헤더 = [
            ("B6", "구분\n(分析)"), ("C6", "ITEM"), ("D6", "MODEL NO."),
            ("E6", "수출\n(USD)"), ("F6", "MOQ\n(起订量)"),
        ]
        우측_헤더 = [
            ("G6", "사진\n(图片)"), ("H6", "이름\n(名字)"),
            ("I6", "바코드\n(条码)"), ("J6", "가격\n(价格)"),
            ("K6", "수량\n(数量)"), ("L6", "합계\n(总额)"),
        ]
        for addr, val in 좌측_헤더:
            c = ws[addr]
            c.value = val
            c.font = _헤더_폰트
            c.fill = _채우기_헤더_좌
            c.alignment = _중앙정렬
            c.border = _테두리()
        for addr, val in 우측_헤더:
            c = ws[addr]
            c.value = val
            c.font = _헤더_폰트
            c.fill = _채우기_헤더_우
            c.alignment = _중앙정렬
            c.border = _테두리()
        ws.row_dimensions[6].height = 54

    # ── 제품 행 ──────────────────────────────────────────────────

    def _제품행_작성(self, ws, 발주: 발주서) -> int:
        행 = 7
        for 제품 in 발주.제품목록:
            ws.row_dimensions[행].height = 122

            data = {
                "B": 제품.번호,
                "C": 제품.item,
                "D": 제품.model_no,
                "E": 제품.수출_usd,
                "F": 제품.moq,
                "H": 제품.이름,
                "I": 제품.바코드,
                "J": 제품.원가,
                "K": 제품.수량,
                "L": 제품.합계,
            }
            for col, val in data.items():
                c = ws[f"{col}{행}"]
                c.value = val
                c.font = _본문_폰트
                c.alignment = _중앙정렬
                c.border = _테두리()

            # G열 테두리 (사진 셀)
            ws[f"G{행}"].border = _테두리()

            # 제품 사진 중앙 삽입 (G열)
            if 제품.사진_경로 and Path(str(제품.사진_경로)).exists():
                _이미지_중앙삽입(
                    ws, str(제품.사진_경로),
                    col_idx=7, row_idx=행,   # G열=7
                    img_w=100, img_h=110,
                    span_cols=1,
                )

            행 += 1

        return 행 - 1

    # ── 포장 이미지 섹션 ─────────────────────────────────────────

    def _포장이미지_삽입(self, ws, 발주: 발주서, 시작행: int):
        if not 발주.포장_이미지들:
            return

        # 섹션 제목 (항목이 있으면 항상 출력)
        ws.merge_cells(f"B{시작행}:L{시작행}")
        ws[f"B{시작행}"] = "포장 요구사항 (从小到大包装要求)"
        ws[f"B{시작행}"].font = Font(name="맑은 고딕", bold=True, size=10)
        ws[f"B{시작행}"].fill = _채우기_포장레이블
        ws[f"B{시작행}"].alignment = _좌측정렬
        ws.row_dimensions[시작행].height = 20

        ws.row_dimensions[시작행 + 1].height = 6

        # 열 그룹: B(2)~D(4), E(5)~G(7), H(8)~J(10) → 3개씩 나란히
        열_그룹   = [(2, 4), (5, 7), (8, 10)]
        IMG_ROW_H = 130   # 이미지 행 높이 (pt)
        TEXT_ROW_H = 50   # 텍스트 행 높이 (pt)

        현재_img_행   = 시작행 + 2
        현재_텍스트_행 = 현재_img_행 + 1

        for i, 항목 in enumerate(발주.포장_이미지들):
            그룹_idx = i % 3

            # 새 행 블록 시작 (4번째 항목부터)
            if 그룹_idx == 0 and i > 0:
                현재_img_행   = 현재_텍스트_행 + 2
                현재_텍스트_행 = 현재_img_행 + 1

            img_행  = 현재_img_행
            txt_행  = img_행 + 1
            현재_텍스트_행 = txt_행

            ws.row_dimensions[img_행].height  = IMG_ROW_H
            ws.row_dimensions[txt_행].height  = TEXT_ROW_H

            시작열, 끝열 = 열_그룹[그룹_idx]
            span = 끝열 - 시작열 + 1

            # 이미지: fit=True → 셀 크기에 맞춰 비율 유지 축소/확대
            경로 = 항목.get("경로")
            if 경로 and Path(str(경로)).exists():
                try:
                    from PIL import Image as _PIL
                    with _PIL.open(str(경로)) as _p:
                        orig_w, orig_h = _p.size
                except Exception:
                    orig_w, orig_h = 300, 200
                _이미지_중앙삽입(
                    ws, str(경로),
                    col_idx=시작열, row_idx=img_행,
                    img_w=orig_w, img_h=orig_h,
                    span_cols=span,
                    fit=True,
                )

            # 텍스트 박스: 노랑 배경 + 흰 글자 bold
            구분 = (항목.get("이름") or "").strip()
            입력 = (항목.get("텍스트") or "").strip()
            if 구분 and 입력:
                텍스트 = f"{구분}, {입력}"
            else:
                텍스트 = 구분 or 입력
            m_s = get_column_letter(시작열)
            m_e = get_column_letter(끝열)
            ws.merge_cells(f"{m_s}{txt_행}:{m_e}{txt_행}")
            c = ws[f"{m_s}{txt_행}"]
            c.value     = 텍스트
            c.font      = Font(name="맑은 고딕", bold=True, size=22, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="FFC000")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── 브랜드별 외박스 표기 (M열 오른쪽, 행2~행6) ───────────────

    def _외박스_섹션(self, ws, 발주: 발주서):
        """
        ILYON / 누아트스튜디오 공용.
        M2:O4 텍스트 + M5:O6 이미지(비율 유지 fit).
        """
        _기본_텍스트 = {
            "ILYON":    "※每个外箱上请填写'ILYON'公司名& 需要蓝色胶带",
            "일리온":   "※每个外箱上请填写'ILYON'公司名& 需要蓝色胶带",
            "누아트스튜디오": "※每个外箱上请填写'NUART'公司名& 需要红色胶带",
        }
        텍스트 = 발주.외박스_텍스트 or _기본_텍스트.get(발주.브랜드, "")

        # M, N, O 열 너비
        for col, w in [("M", 18), ("N", 18), ("O", 18)]:
            ws.column_dimensions[col].width = w

        # ── 텍스트 영역: M2:O4 ───────────────────────────────────
        ws.merge_cells("M2:O4")
        c = ws["M2"]
        c.value     = 텍스트
        c.font      = Font(name="맑은 고딕", bold=True, size=10, color="CC0000")
        c.fill      = PatternFill("solid", fgColor="FFEB9C")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _테두리()
        for addr in ("N2", "O2", "M3", "N3", "O3", "M4", "N4", "O4"):
            ws[addr].border = _테두리()
        for r in (2, 3, 4):
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 26)

        # ── 이미지 영역: M5:O6 (비율 유지 fit) ───────────────────
        if not (발주.외박스_이미지 and Path(str(발주.외박스_이미지)).exists()):
            return

        ws.merge_cells("M5:O6")
        ws["M5"].border = _테두리()
        for addr in ("N5", "O5", "M6", "N6", "O6"):
            ws[addr].border = _테두리()
        ws.row_dimensions[5].height = max(ws.row_dimensions[5].height or 0, 70)
        ws.row_dimensions[6].height = max(ws.row_dimensions[6].height or 0, 70)

        # 원본 이미지 크기 읽기
        try:
            from PIL import Image as PILImage
            with PILImage.open(str(발주.외박스_이미지)) as pil:
                orig_w, orig_h = pil.size
        except Exception:
            orig_w, orig_h = 200, 150  # 폴백

        _이미지_중앙삽입(
            ws, str(발주.외박스_이미지),
            col_idx=13, row_idx=5,   # M열=13
            img_w=orig_w, img_h=orig_h,
            span_cols=3, span_rows=2,
            fit=True,
        )
