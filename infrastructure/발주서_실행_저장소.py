"""
발주서_실행_저장소 - 원본 발주서에 수량을 채워 발주서 폴더에 저장
"""
import shutil
from pathlib import Path
import openpyxl


class 발주서_실행_저장소:

    def __init__(self, 발주서_디렉토리: Path):
        self._발주서_디렉토리 = 발주서_디렉토리

    def 발주서_생성(
        self,
        원본_경로: Path,
        출력_경로: Path,
        바코드_수량: dict[str, int],
        예상_도착일: str = "",
    ) -> Path:
        """
        원본 발주서를 복사하고 바코드에 해당하는 수량을 채워 저장.
        열 인덱스: B=2, I=9(바코드), J=10(원가), K=11(수량), L=12(합계)
        """
        출력_경로.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(원본_경로, 출력_경로)

        wb = openpyxl.load_workbook(출력_경로)
        ws = wb.active

        if 예상_도착일:
            ws["G4"] = 예상_도착일

        # 데이터 행 범위 확정 (번호열 B가 비면 종료)
        마지막_행 = 6
        for row_idx in range(7, ws.max_row + 1):
            if not ws.cell(row=row_idx, column=2).value:
                break
            마지막_행 = row_idx

        for row_idx in range(7, 마지막_행 + 1):
            # I열 값이 '한국상품명_바코드번호' 형식일 수 있으므로 파서와 동일하게 분리
            raw = str(ws.cell(row=row_idx, column=9).value or "").strip()
            바코드_str = self._바코드_추출(raw)

            if 바코드_str in 바코드_수량:
                원가 = float(ws.cell(row=row_idx, column=10).value or 0)
                수량 = 바코드_수량[바코드_str]
                ws.cell(row=row_idx, column=11).value = 수량
                ws.cell(row=row_idx, column=12).value = round(원가 * 수량, 2)
            else:
                # 행 내용 비우고 높이를 0으로 → 이미지 앵커 유지하면서 시각적으로 숨김
                for col in range(1, (ws.max_column or 12) + 1):
                    ws.cell(row=row_idx, column=col).value = None
                ws.row_dimensions[row_idx].height = 0.1

        wb.save(출력_경로)
        return 출력_경로

    @staticmethod
    def _바코드_추출(raw: str) -> str:
        """'한국상품명_바코드번호' → '바코드번호', 바코드단독 → 그대로 반환."""
        if "_" in raw:
            후보 = raw.rsplit("_", 1)[-1].strip()
            if 후보.isdigit():
                return 후보
        return raw

    def 출력_경로_계산(self, 발주_유형: str, po코드: str) -> Path:
        """'정기발주', 'NF(695)' → 발주서_DIR/정기발주/NF(695).xlsx"""
        return self._발주서_디렉토리 / 발주_유형 / f"{po코드}.xlsx"
