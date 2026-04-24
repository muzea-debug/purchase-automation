import openpyxl
import pytest
from pathlib import Path
from infrastructure.발주서_실행_저장소 import 발주서_실행_저장소


@pytest.fixture
def 원본_발주서(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["G4"] = ""
    ws["B7"] = 1
    ws["H7"] = "테스트 어댑터"
    ws["I7"] = "8800000000001"
    ws["J7"] = 100.0
    ws["K7"] = 0
    ws["L7"] = 0
    ws["B8"] = 2
    ws["H8"] = "테스트 허브"
    ws["I8"] = "8800000000002"
    ws["J8"] = 70.0
    ws["K8"] = 0
    ws["L8"] = 0
    파일 = tmp_path / "원본.xlsx"
    wb.save(파일)
    return 파일


def test_수량_채우기(원본_발주서, tmp_path):
    저장소 = 발주서_실행_저장소(tmp_path)
    출력 = tmp_path / "정기발주" / "NF(695).xlsx"
    출력.parent.mkdir(parents=True, exist_ok=True)

    저장소.발주서_생성(
        원본_경로=원본_발주서,
        출력_경로=출력,
        바코드_수량={"8800000000001": 500},
        한국_요청_도착일="2026. 4. 30",
    )

    wb = openpyxl.load_workbook(출력)
    ws = wb.active
    assert ws["K7"].value == 500
    assert ws["L7"].value == pytest.approx(50000.0)
    assert ws["K8"].value == 0
    assert ws["G4"].value == "2026. 4. 30"


def test_미포함_바코드_수량_유지(원본_발주서, tmp_path):
    저장소 = 발주서_실행_저장소(tmp_path)
    출력 = tmp_path / "신기종" / "APn(5).xlsx"
    출력.parent.mkdir(parents=True, exist_ok=True)

    저장소.발주서_생성(
        원본_경로=원본_발주서,
        출력_경로=출력,
        바코드_수량={"8800000000002": 200},
    )
    wb = openpyxl.load_workbook(출력)
    ws = wb.active
    assert ws["K8"].value == 200
    assert ws["L8"].value == pytest.approx(14000.0)
