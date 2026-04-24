"""발주서원본_파서 테스트"""
import openpyxl
from pathlib import Path
from infrastructure.발주서원본_파서 import 발주서원본_파서


def _샘플_xlsx(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B7"] = 1
    ws["C7"] = "Test Adapter"
    ws["D7"] = "MODEL-001"
    ws["E7"] = 5.12
    ws["F7"] = "1000"
    ws["H7"] = "누아트 테스트 어댑터, 블랙"
    ws["I7"] = "8800000000001"
    ws["J7"] = 100.0
    ws["K7"] = 0
    ws["L7"] = 0
    ws["B8"] = 2
    ws["C8"] = "USB Hub"
    ws["D8"] = "MODEL-002"
    ws["E8"] = 3.50
    ws["F8"] = "500"
    ws["H8"] = "누아트 USB 허브, 화이트"
    ws["I8"] = "8800000000002"
    ws["J8"] = 70.0
    ws["K8"] = 0
    ws["L8"] = 0
    파일 = tmp_path / "富坤电器_NUART 테스트 제품(NUART전용).xlsx"
    wb.save(파일)
    return tmp_path


def test_카탈로그_로드_바코드_포함(tmp_path):
    _샘플_xlsx(tmp_path)
    카탈로그 = 발주서원본_파서(tmp_path).카탈로그_로드()
    assert "8800000000001" in 카탈로그
    assert "8800000000002" in 카탈로그


def test_카탈로그_상품_정보_정확도(tmp_path):
    _샘플_xlsx(tmp_path)
    상품 = 발주서원본_파서(tmp_path).카탈로그_로드()["8800000000001"]
    assert 상품.이름 == "누아트 테스트 어댑터, 블랙"
    assert 상품.브랜드 == "NUART"
    assert 상품.제조사 == "富坤电器"
    assert 상품.원가 == 100.0


def test_빈_디렉토리_빈_카탈로그(tmp_path):
    assert 발주서원본_파서(tmp_path).카탈로그_로드() == {}
