"""
재고현황 라우트 - 파일 업로드 / 분석 / 최근 결과 불러오기 API
"""
import uuid
import json
import shutil
from pathlib import Path
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

import config
from application.재고현황_유스케이스 import 재고현황_유스케이스
from infrastructure.google_sheets_저장소 import GoogleSheets저장소
from infrastructure.엑셀_시트_저장소 import 엑셀시트저장소

bp_재고현황 = Blueprint("재고현황", __name__)


def _원본발주서_업데이트(바코드_set: set, 필드들: dict) -> int:
    """
    통합발주서원본_DIR 내 xlsx 파일들에서 바코드 일치 행의 MOQ·원가를 업데이트.
    반환: 수정된 파일 수
    """
    moq_val  = 필드들.get("moq")
    원가_val = 필드들.get("원가_위안")
    if moq_val is None and 원가_val is None:
        return 0  # 업데이트할 필드 없음

    import openpyxl, re as _re

    def _bc_from_cell(v) -> str:
        """I열 값 → 바코드 추출 ('한국상품명_바코드' 또는 '바코드단독')"""
        raw = str(v).strip()
        if "_" in raw:
            후보 = raw.rsplit("_", 1)[-1].strip()
            if 후보.isdigit():
                return 후보
        return raw

    수정파일수 = 0
    try:
        dir_path = config.통합발주서원본_DIR
        if not dir_path.exists():
            return 0
        for xlsx in dir_path.glob("*.xlsx"):
            if xlsx.name.startswith("~$"):
                continue
            try:
                wb = openpyxl.load_workbook(xlsx)
                ws = wb.active
                changed = False
                for row in ws.iter_rows(min_row=7):
                    if len(row) < 10:
                        break
                    bc_cell = row[8]  # I열 (0-indexed)
                    if not bc_cell.value:
                        break
                    bc = _bc_from_cell(bc_cell.value)
                    if bc not in 바코드_set:
                        continue
                    if moq_val is not None:
                        row[5].value = moq_val   # F열
                        changed = True
                    if 원가_val is not None:
                        row[9].value = float(원가_val)  # J열
                        changed = True
                if changed:
                    wb.save(xlsx)
                    수정파일수 += 1
                wb.close()
            except Exception as e:
                print(f"[원본발주서_업데이트] {xlsx.name} 실패: {e}")
    except Exception as e:
        print(f"[원본발주서_업데이트] 오류: {e}")
    return 수정파일수


# ── 메모리 캐시 ───────────────────────────────────────────────────
_재고현황_캐시: dict | None = None


def 재고현황_캐시_로드():
    """저장된 JSON을 서버 메모리에 미리 로드."""
    global _재고현황_캐시
    try:
        if config.재고현황_저장_파일.exists():
            with open(config.재고현황_저장_파일, encoding="utf-8") as f:
                _재고현황_캐시 = json.load(f)
            print(f"[재고현황] 캐시 로드 완료: {len(_재고현황_캐시.get('rows', []))}행", flush=True)
    except Exception as e:
        print(f"[재고현황] 캐시 로드 실패: {e}", flush=True)


def _캐시_갱신(data: dict):
    global _재고현황_캐시
    _재고현황_캐시 = data


_아카이브_DIR = config.재고현황_저장_파일.parent / "재고현황_아카이브"

def _아카이브_저장(결과: dict):
    """날짜별 스냅샷을 아카이브 폴더에 저장. 하루 1회만 (같은 날 덮어씀)."""
    try:
        _아카이브_DIR.mkdir(parents=True, exist_ok=True)
        오늘 = datetime.now().strftime("%Y-%m-%d")
        아카이브_파일 = _아카이브_DIR / f"{오늘}.json"
        # 핵심 컬럼만 저장 (파일 경량화)
        _저장_키 = {"바코드", "상품명", "브랜드", "대분류", "소분류", "시즌구분",
                    "재고합계", "쿠팡주간판매", "일반채널판매", "발주량산출",
                    "재고상태", "중국재고_쿠팡", "중국재고_일반", "사방넷재고",
                    "중국쿠팡_미입고", "중국일반_미입고", "쿠팡발주수량"}
        슬림_rows = [{k: r.get(k) for k in _저장_키} for r in 결과.get("rows", [])]
        with open(아카이브_파일, "w", encoding="utf-8") as f:
            json.dump({"날짜": 오늘, "rows": 슬림_rows}, f, ensure_ascii=False)
        print(f"[아카이브] {오늘} 저장 완료 ({len(슬림_rows)}개 제품)")
    except Exception as e:
        print(f"[아카이브] 저장 실패: {e}")


def _sheets_연결():
    """Google Sheets 우선, 없으면 로컬 Excel 파일 사용."""
    if config.GOOGLE_SERVICE_ACCOUNT_JSON and config.GOOGLE_SHEET_ID:
        try:
            return GoogleSheets저장소(
                config.GOOGLE_SERVICE_ACCOUNT_JSON,
                config.GOOGLE_SHEET_ID,
                config.GOOGLE_SHEET_NAME,
            )
        except Exception as e:
            print(f"[재고현황] Sheets 연결 실패: {e}")

    if config.로컬시트_파일.exists():
        return 엑셀시트저장소(config.로컬시트_파일)

    return None


@bp_재고현황.route("/api/재고현황/로컬시트_초기화", methods=["POST"])
def 로컬시트_초기화_api():
    """
    구글 시트 대체용 Excel 파일 초기 생성.
    body (multipart, 모두 선택):
      엔대시소스, 주문프로그램, 재고현황표
    """
    from infrastructure.로컬시트_초기화 import 로컬시트_초기화
    import uuid, shutil
    from werkzeug.utils import secure_filename

    덮어쓰기 = request.form.get("덮어쓰기", "false").lower() == "true"

    temp_dir = config.재고현황_UPLOAD_DIR / uuid.uuid4().hex[:8]
    temp_dir.mkdir(parents=True, exist_ok=True)

    def _저장(field):
        f = request.files.get(field)
        if not f or not f.filename:
            return None
        p = temp_dir / secure_filename(f.filename)
        f.save(str(p))
        return p

    엔대시소스 = _저장("엔대시소스") or (
        config.로컬시트_엔대시소스 if config.로컬시트_엔대시소스.exists() else None
    )
    주문프로그램 = _저장("주문프로그램") or (
        config.로컬시트_주문프로그램 if config.로컬시트_주문프로그램.exists() else None
    )
    재고현황표 = _저장("재고현황표") or (
        config.로컬시트_재고현황표 if config.로컬시트_재고현황표.exists() else None
    )

    try:
        생성됨 = 로컬시트_초기화(
            저장_경로=config.로컬시트_파일,
            엔대시_소스=엔대시소스,
            주문프로그램=주문프로그램,
            재고현황표=재고현황표,
            덮어쓰기=덮어쓰기,
        )
        sheets_ok = config.로컬시트_파일.exists()
        return jsonify({
            "success": True,
            "생성됨": 생성됨,
            "파일": str(config.로컬시트_파일),
            "sheets_연결": sheets_ok,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@bp_재고현황.route("/dashboard")
def 재고현황_대시보드():
    return render_template("dashboard.html")


@bp_재고현황.route("/재고현황")
def 재고현황_페이지():
    구글_ok = bool(config.GOOGLE_SERVICE_ACCOUNT_JSON and config.GOOGLE_SHEET_ID)
    로컬_ok = config.로컬시트_파일.exists()
    sheets_연결 = 구글_ok or 로컬_ok
    sheets_유형 = "google" if 구글_ok else ("local" if 로컬_ok else "none")

    # 캐시 없으면 즉시 로드 (첫 번째 접근 시)
    if _재고현황_캐시 is None:
        재고현황_캐시_로드()

    return render_template("재고현황.html",
                           sheets_연결=sheets_연결,
                           sheets_유형=sheets_유형)


@bp_재고현황.route("/api/재고현황/데이터", methods=["GET"])
def 재고현황_데이터():
    """메모리 캐시 즉시 반환 (HTML 임베드 대체)."""
    if _재고현황_캐시 is None:
        재고현황_캐시_로드()
    if not _재고현황_캐시:
        return jsonify({"success": False, "rows": []})
    return jsonify(_재고현황_캐시)


@bp_재고현황.route("/api/재고현황/최근", methods=["GET"])
def 재고현황_최근():
    """마지막으로 저장된 분석 결과 반환."""
    저장_파일 = config.재고현황_저장_파일
    if not 저장_파일.exists():
        return jsonify({"success": False, "error": "저장된 데이터가 없습니다. 소스파일을 업데이트해주세요."})
    try:
        with open(저장_파일, encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp_재고현황.route("/api/재고현황/시트디버그", methods=["GET"])
def 재고현황_시트디버그():
    """시트에서 실제로 읽히는 데이터 확인용 (테스트 전용)."""
    try:
        sheets = _sheets_연결()
        if not sheets:
            return jsonify({"error": "sheets 연결 안됨"})
        헤더, 행들 = sheets.발주요청_읽기(config.ENDASH_발주요청_시트)
        return jsonify({
            "시트명": config.ENDASH_발주요청_시트,
            "헤더": 헤더,
            "총행수": len(행들),
            "샘플_5행": 행들[:5],
        })
    except Exception as e:
        return jsonify({"error": str(e)})


@bp_재고현황.route("/api/재고현황/시트업데이트", methods=["POST"])
def 재고현황_시트업데이트():
    """저장된 기존 데이터를 유지하면서 구글시트만 다시 읽어 미입고·중국재고 갱신."""
    try:
        기존_행맵: dict = {}
        if config.재고현황_저장_파일.exists():
            with open(config.재고현황_저장_파일, encoding="utf-8") as f:
                기존 = json.load(f)
            기존_행맵 = {r["바코드"]: r for r in 기존.get("rows", []) if r.get("바코드")}

        if not 기존_행맵:
            return jsonify({"success": False, "error": "저장된 데이터가 없습니다. 먼저 전체 조회를 실행해 주세요."}), 400

        uc = 재고현황_유스케이스(sheets_저장소=_sheets_연결())
        결과 = uc.분석(기존_행맵=기존_행맵)  # 파일 없이 시트만 읽기

        if 결과.get("success"):
            결과["업데이트_시각"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            config.재고현황_저장_파일.parent.mkdir(parents=True, exist_ok=True)
            with open(config.재고현황_저장_파일, "w", encoding="utf-8") as f:
                json.dump(결과, f, ensure_ascii=False)
            _캐시_갱신(결과)

        return jsonify(결과)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/셀_저장", methods=["POST"])
def 재고현황_셀_저장():
    """
    실재고·비고 등 수동 입력값 저장 및 브로드캐스트용 타임스탬프 갱신.
    body: {"바코드": "...", "필드": "실재고", "값": 100}
    수동 입력 가능 필드: 실재고, 비고, 쿠팡조정, 쿠팡조정사유, 일반채널조정, 일반채널사유, 최종발주량
    """
    _수동필드 = {"물류실재고", "구매팀발주", "쿠팡조정", "쿠팡의견", "일반조정", "일반의견", "비고", "승인", "물류확인"}
    try:
        data = request.get_json(force=True)
        바코드 = str(data.get("바코드", "")).strip()
        필드   = str(data.get("필드", "")).strip()
        값     = data.get("값", "")

        if not 바코드 or 필드 not in _수동필드:
            return jsonify({"success": False, "error": "바코드 또는 필드 오류"}), 400

        저장_파일 = config.재고현황_저장_파일
        if not 저장_파일.exists():
            return jsonify({"success": False, "error": "저장된 데이터 없음"}), 404

        with open(저장_파일, encoding="utf-8") as f:
            저장_data = json.load(f)

        rows = 저장_data.get("rows", [])
        갱신됨 = False
        for row in rows:
            if row.get("바코드") == 바코드:
                row[필드] = 값
                # 실재고 변경 시 재고합계 재계산
                if 필드 == "실재고":
                    try:
                        실재고_n = float(값) if 값 not in ("", None) else 0.0
                        row["실재고"] = 실재고_n
                        row["재고합계"] = round(
                            float(row.get("중국재고_쿠팡", 0))
                            + float(row.get("중국재고_일반", 0))
                            + float(row.get("사방넷재고", 0))
                            + float(row.get("중국쿠팡_미입고", 0))
                            + float(row.get("중국일반_미입고", 0))
                            + 실재고_n, 0
                        )
                    except (TypeError, ValueError):
                        pass
                갱신됨 = True
                break

        if not 갱신됨:
            return jsonify({"success": False, "error": "바코드 없음"}), 404

        저장_data["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(저장_파일, "w", encoding="utf-8") as f:
            json.dump(저장_data, f, ensure_ascii=False)
        _캐시_갱신(저장_data)

        return jsonify({"success": True, "last_modified": 저장_data["last_modified"]})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/제품DB_수정", methods=["POST"])
def 재고현황_제품DB_수정():
    """
    제품DB 역방향 업데이트 (배치 저장).
    body: {"바코드": "...", "필드들": {"moq":"15~20", "기본_lt":45, ...}}
    수정 가능 필드: moq, 기본_lt, 원가_위안, 최종원가, 발주서명, 안전일수, 주의사항, 브랜드
    """
    _허용_필드 = {"moq", "기본_lt", "원가_위안", "최종원가", "발주서명", "안전일수", "주의사항", "브랜드"}
    try:
        data   = request.get_json(force=True)
        바코드 = str(data.get("바코드", "")).strip()
        필드들 = {k: v for k, v in data.get("필드들", {}).items() if k in _허용_필드}

        if not 바코드 or not 필드들:
            return jsonify({"success": False, "error": "바코드 또는 필드들 없음"}), 400

        저장소 = _sheets_연결()
        ok = 저장소.제품DB_수정(바코드, 필드들)
        if not ok:
            return jsonify({"success": False, "error": "바코드 없음 또는 수정 실패"}), 404

        # 메모리 캐시 반영 (재고현황 JSON rows)
        if _재고현황_캐시:
            for row in _재고현황_캐시.get("rows", []):
                if row.get("바코드") == 바코드:
                    for k, v in 필드들.items():
                        row[k] = v
                    if "moq" in 필드들:
                        import re as _re
                        nums = _re.findall(r'\d+', str(필드들["moq"]))
                        row["발주확정용_moq"] = float(max(int(n) for n in nums)) if nums else 0.0
                    if "기본_lt" in 필드들:
                        try:
                            import re as _re
                            nums = _re.findall(r'\d+', str(필드들["기본_lt"]))
                            row["기본_lt"] = float(max(int(n) for n in nums)) if nums else 0.0
                        except Exception:
                            pass
                    break
            try:
                config.재고현황_저장_파일.parent.mkdir(parents=True, exist_ok=True)
                with open(config.재고현황_저장_파일, "w", encoding="utf-8") as _f:
                    json.dump(_재고현황_캐시, _f, ensure_ascii=False)
            except Exception as _e:
                print(f"[단건수정] JSON 파일 저장 실패: {_e}")

        _원본발주서_업데이트({바코드}, 필드들)
        return jsonify({"success": True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/제품DB_일괄수정", methods=["POST"])
def 재고현황_제품DB_일괄수정():
    """
    body: {"바코드들": ["...", ...], "필드들": {"기본_lt": 45, ...}}
    수정 가능 필드: moq, 기본_lt, 원가_위안, 최종원가, 발주서명, 안전일수, 주의사항, 브랜드
    """
    _허용_필드 = {"moq", "기본_lt", "원가_위안", "최종원가", "발주서명", "안전일수", "주의사항", "브랜드"}
    import re as _re
    try:
        data = request.get_json(force=True)
        바코드들 = [str(b).strip() for b in data.get("바코드들", []) if b]
        필드들 = {k: v for k, v in data.get("필드들", {}).items() if k in _허용_필드}

        if not 바코드들 or not 필드들:
            return jsonify({"success": False, "error": "바코드들 또는 필드들 없음"}), 400

        저장소 = _sheets_연결()
        수정된수 = 0
        for 바코드 in 바코드들:
            try:
                ok = 저장소.제품DB_수정(바코드, 필드들)
                if ok:
                    수정된수 += 1
            except Exception:
                pass

        # 메모리 캐시 + JSON 파일 반영
        if _재고현황_캐시:
            bc_set = set(바코드들)
            for row in _재고현황_캐시.get("rows", []):
                if row.get("바코드") in bc_set:
                    for k, v in 필드들.items():
                        row[k] = v
                    if "moq" in 필드들:
                        nums = _re.findall(r'\d+', str(필드들["moq"]))
                        row["발주확정용_moq"] = float(max(int(n) for n in nums)) if nums else 0.0
                    if "기본_lt" in 필드들:
                        nums = _re.findall(r'\d+', str(필드들["기본_lt"]))
                        row["기본_lt"] = float(max(int(n) for n in nums)) if nums else 0.0
            try:
                config.재고현황_저장_파일.parent.mkdir(parents=True, exist_ok=True)
                with open(config.재고현황_저장_파일, "w", encoding="utf-8") as _f:
                    json.dump(_재고현황_캐시, _f, ensure_ascii=False)
            except Exception as _e:
                print(f"[일괄수정] JSON 파일 저장 실패: {_e}")

        _원본발주서_업데이트(set(바코드들), 필드들)
        return jsonify({"success": True, "수정된수": 수정된수})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/전체승인", methods=["POST"])
def 재고현황_전체승인():
    """body: {"바코드들": ["...", ...], "승인값": true/false}"""
    try:
        data = request.get_json(force=True)
        바코드들 = set(str(b) for b in data.get("바코드들", []))
        승인값   = bool(data.get("승인값", True))

        저장_파일 = config.재고현황_저장_파일
        if not 저장_파일.exists():
            return jsonify({"success": False, "error": "저장된 데이터 없음"}), 404

        with open(저장_파일, encoding="utf-8") as f:
            저장_data = json.load(f)

        for row in 저장_data.get("rows", []):
            if row.get("바코드") in 바코드들:
                row["승인"] = 승인값

        저장_data["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(저장_파일, "w", encoding="utf-8") as f:
            json.dump(저장_data, f, ensure_ascii=False)
        _캐시_갱신(저장_data)

        return jsonify({"success": True, "last_modified": 저장_data["last_modified"]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/아카이브_목록", methods=["GET"])
def 재고현황_아카이브_목록():
    """저장된 JSON 아카이브 날짜 목록 반환."""
    if not _아카이브_DIR.exists():
        return jsonify({"success": True, "날짜들": []})
    날짜들 = sorted(
        [f.stem for f in _아카이브_DIR.glob("*.json")],
        reverse=True
    )
    return jsonify({"success": True, "날짜들": 날짜들})


@bp_재고현황.route("/api/재고현황/소스_아카이브_목록", methods=["GET"])
def 재고현황_소스_아카이브_목록():
    """재고현황_아카이브/ 소스 파일 폴더 목록 반환."""
    from infrastructure.아카이브_스캐너 import 아카이브_스캐너
    목록 = 아카이브_스캐너().날짜_목록()
    return jsonify({"success": True, "목록": 목록})


@bp_재고현황.route("/api/재고현황/일괄_아카이브_분석", methods=["POST"])
def 재고현황_일괄_아카이브_분석():
    """
    소스 아카이브 폴더 전체를 순서대로 분석 → 각각 스냅샷 저장.
    body: {"덮어쓰기": false}  — false면 이미 스냅샷 있는 날짜 건너뜀
    Returns: {"결과": [{"날짜":..., "행수":..., "skipped":bool}, ...]}
    """
    from infrastructure.아카이브_스캐너 import 아카이브_스캐너
    try:
        data = request.get_json(force=True) or {}
        덮어쓰기 = data.get("덮어쓰기", False)

        목록 = 아카이브_스캐너().날짜_목록()
        if not 목록:
            return jsonify({"success": True, "결과": [], "message": "소스 폴더 없음"})

        sheets = _sheets_연결()
        uc = 재고현황_유스케이스(sheets_저장소=sheets)

        _저장_키 = {"바코드", "상품명", "브랜드", "대분류", "소분류", "시즌구분",
                    "재고합계", "쿠팡주간판매", "일반채널판매", "발주량산출",
                    "재고상태", "중국재고_쿠팡", "중국재고_일반", "사방넷재고",
                    "중국쿠팡_미입고", "중국일반_미입고", "쿠팡발주수량"}

        결과_목록 = []
        _아카이브_DIR.mkdir(parents=True, exist_ok=True)

        for item in 목록:
            폴더명 = item["폴더명"]
            날짜   = item["날짜"]
            아카이브_파일 = _아카이브_DIR / f"{날짜}.json"

            if not 덮어쓰기 and 아카이브_파일.exists():
                결과_목록.append({"날짜": 날짜, "행수": None, "skipped": True, "reason": "이미 존재"})
                continue

            try:
                분석결과 = uc.분석_아카이브(폴더명)
                if 분석결과.get("success"):
                    슬림_rows = [{k: r.get(k) for k in _저장_키} for r in 분석결과.get("rows", [])]
                    with open(아카이브_파일, "w", encoding="utf-8") as f:
                        json.dump({"날짜": 날짜, "rows": 슬림_rows}, f, ensure_ascii=False)
                    결과_목록.append({"날짜": 날짜, "행수": len(슬림_rows), "skipped": False})
                else:
                    결과_목록.append({"날짜": 날짜, "행수": 0, "skipped": False, "error": 분석결과.get("error", "실패")})
            except Exception as e:
                결과_목록.append({"날짜": 날짜, "skipped": False, "error": str(e)})

        완료 = sum(1 for r in 결과_목록 if not r["skipped"] and not r.get("error"))
        건너뜀 = sum(1 for r in 결과_목록 if r["skipped"])
        return jsonify({"success": True, "결과": 결과_목록, "완료": 완료, "건너뜀": 건너뜀})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/과거_분석", methods=["POST"])
def 재고현황_과거_분석():
    """
    body: {"폴더명": "2026.04.08"}
    → 아카이브 소스 파일로 해당 시점 재고현황 분석 후 JSON 스냅샷 저장.
    """
    try:
        data = request.get_json(force=True)
        폴더명 = (data.get("폴더명") or "").strip()
        if not 폴더명:
            return jsonify({"success": False, "error": "폴더명 필요"}), 400

        # 제품 마스터는 현재 sheets 연결 (과거 파일에는 마스터 없음)
        uc = 재고현황_유스케이스(sheets_저장소=_sheets_연결())
        결과 = uc.분석_아카이브(폴더명)

        if 결과.get("success"):
            # 날짜 추출 (YYYY.MM.DD → YYYY-MM-DD)
            parts = 폴더명.split(".")
            날짜 = "-".join(parts) if len(parts) == 3 else 폴더명
            결과["업데이트_시각"] = 날짜

            # JSON 스냅샷 저장
            _저장_키 = {"바코드", "상품명", "브랜드", "대분류", "소분류", "시즌구분",
                        "재고합계", "쿠팡주간판매", "일반채널판매", "발주량산출",
                        "재고상태", "중국재고_쿠팡", "중국재고_일반", "사방넷재고",
                        "중국쿠팡_미입고", "중국일반_미입고", "쿠팡발주수량"}
            슬림_rows = [{k: r.get(k) for k in _저장_키} for r in 결과.get("rows", [])]
            _아카이브_DIR.mkdir(parents=True, exist_ok=True)
            아카이브_파일 = _아카이브_DIR / f"{날짜}.json"
            with open(아카이브_파일, "w", encoding="utf-8") as f:
                json.dump({"날짜": 날짜, "rows": 슬림_rows}, f, ensure_ascii=False)
            결과["스냅샷_저장"] = str(아카이브_파일)

        return jsonify(결과)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp_재고현황.route("/api/재고현황/바코드_이력", methods=["GET"])
def 재고현황_바코드_이력():
    """
    ?barcodes=A,B,C
    각 바코드의 스냅샷별 시계열 반환.
    {바코드: [{날짜, 재고합계, 쿠팡주간판매, 일반채널판매, 발주량산출, 재고소진일}, ...]}
    """
    barcodes_raw = request.args.get("barcodes", "")
    barcodes = [b.strip() for b in barcodes_raw.split(",") if b.strip()]
    if not barcodes:
        return jsonify({"success": False, "error": "barcodes 필요"}), 400

    bc_set = set(barcodes)
    결과: dict[str, list] = {bc: [] for bc in barcodes}

    if _아카이브_DIR.exists():
        for 파일 in sorted(_아카이브_DIR.glob("*.json")):
            날짜 = 파일.stem
            try:
                with open(파일, encoding="utf-8") as f:
                    snap = json.load(f)
                for row in snap.get("rows", []):
                    bc = str(row.get("바코드", ""))
                    if bc in bc_set:
                        일판매 = (float(row.get("쿠팡주간판매") or 0) + float(row.get("일반채널판매") or 0)) / 7
                        재고 = float(row.get("재고합계") or 0)
                        소진일 = round(재고 / 일판매) if 일판매 > 0 else None
                        결과[bc].append({
                            "날짜": 날짜,
                            "재고합계": 재고,
                            "쿠팡주간판매": float(row.get("쿠팡주간판매") or 0),
                            "일반채널판매": float(row.get("일반채널판매") or 0),
                            "발주량산출": float(row.get("발주량산출") or 0),
                            "재고소진일": 소진일,
                        })
            except Exception:
                pass

    return jsonify({"success": True, "이력": 결과})


@bp_재고현황.route("/api/재고현황/아카이브/<date_str>", methods=["GET"])
def 재고현황_아카이브_조회(date_str):
    """특정 날짜 아카이브 데이터 반환. date_str 형식: YYYY-MM-DD"""
    if "/" in date_str or ".." in date_str or not date_str.replace("-", "").isdigit():
        return jsonify({"success": False, "error": "잘못된 날짜"}), 400
    파일 = _아카이브_DIR / f"{date_str}.json"
    if not 파일.exists():
        return jsonify({"success": False, "error": "해당 날짜 데이터 없음"}), 404
    with open(파일, encoding="utf-8") as f:
        data = json.load(f)
    return jsonify({"success": True, **data})


@bp_재고현황.route("/api/재고현황/엑셀_다운로드", methods=["GET"])
def 재고현황_엑셀_다운로드():
    """현재 저장된 재고현황 데이터를 템플릿 기반 xlsx로 다운로드."""
    import io
    from copy import copy
    import openpyxl

    저장_파일 = config.재고현황_저장_파일
    if not 저장_파일.exists():
        return jsonify({"success": False, "error": "저장된 데이터가 없습니다"}), 404

    with open(저장_파일, encoding="utf-8") as f:
        data = json.load(f)
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"success": False, "error": "데이터가 없습니다"}), 404

    if not config.재고현황_템플릿.exists():
        return jsonify({"success": False, "error": "재고현황 템플릿 파일이 없습니다"}), 404

    wb = openpyxl.load_workbook(config.재고현황_템플릿)
    ws = wb.active
    max_col = ws.max_column

    # 데이터 행 스타일 캡처 (홀수행=row5, 짝수행=row6) — 삭제 전에 수행
    def _capture_style(row_num):
        styles = {}
        for c in range(1, max_col + 1):
            cell = ws.cell(row_num, c)
            styles[c] = {
                "font":      copy(cell.font),
                "fill":      copy(cell.fill),
                "border":    copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }
        return styles

    style_odd  = _capture_style(5) if ws.max_row >= 5 else {}
    style_even = _capture_style(6) if ws.max_row >= 6 else style_odd

    # 기존 데이터 행(5행~) 삭제
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)

    def _v(row, key, default=""):
        v = row.get(key, default)
        return v if v is not None else default

    COL_MAP = {
        1:  lambda r: _v(r, "시즌구분"),
        2:  lambda r: _v(r, "대분류"),
        3:  lambda r: _v(r, "소분류"),
        4:  lambda r: _v(r, "바코드"),
        5:  lambda r: _v(r, "상품명"),
        6:  lambda r: _v(r, "브랜드"),
        7:  lambda r: _v(r, "발주서명"),
        8:  lambda r: _v(r, "상품명"),
        9:  lambda r: _v(r, "주의사항"),
        10: lambda r: _v(r, "원가_위안"),
        13: lambda r: _v(r, "최종원가"),
        14: lambda r: _v(r, "기본_lt"),
        15: lambda r: _v(r, "moq"),
        20: lambda r: _v(r, "재고합계"),
        21: lambda r: _v(r, "예상판매_lt"),
        22: lambda r: _v(r, "발주량산출"),
        24: lambda r: _v(r, "안전재고일수"),
        25: lambda r: _v(r, "구매팀발주"),
        26: lambda r: _v(r, "비고"),
        27: lambda r: _v(r, "쿠팡조정"),
        28: lambda r: _v(r, "쿠팡의견"),
        29: lambda r: _v(r, "일반조정"),
        30: lambda r: _v(r, "일반의견"),
        32: lambda r: _v(r, "중국재고_쿠팡"),
        33: lambda r: _v(r, "중국재고_일반"),
        34: lambda r: _v(r, "재고상태"),
        36: lambda r: _v(r, "쿠팡fc재고"),
        37: lambda r: _v(r, "입고예정_코드"),
        38: lambda r: _v(r, "입고일자"),
        40: lambda r: _v(r, "중국쿠팡_미입고"),
        41: lambda r: _v(r, "중국일반_미입고"),
        42: lambda r: _v(r, "재고합계"),
        43: lambda r: _v(r, "쿠팡발주수량"),
        44: lambda r: _v(r, "쿠팡주간판매"),
        45: lambda r: round(float(_v(r, "쿠팡주간판매") or 0) * 6, 1),
        46: lambda r: _v(r, "일반채널판매"),
        47: lambda r: round(float(_v(r, "일반채널판매") or 0) * 6, 1),
        48: lambda r: _v(r, "발주량산출"),
    }

    for i, row in enumerate(rows):
        row_idx = i + 5
        style = style_odd if i % 2 == 0 else style_even
        for c in range(1, max_col + 1):
            cell = ws.cell(row_idx, c)
            s = style.get(c, {})
            if s.get("font"):      cell.font      = copy(s["font"])
            if s.get("fill"):      cell.fill      = copy(s["fill"])
            if s.get("border"):    cell.border    = copy(s["border"])
            if s.get("alignment"): cell.alignment = copy(s["alignment"])
            if s.get("number_format"): cell.number_format = s["number_format"]
        for col, fn in COL_MAP.items():
            try:
                ws.cell(row_idx, col).value = fn(row)
            except Exception:
                pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    오늘 = datetime.now().strftime("%Y%m%d")
    return send_file(
        out,
        as_attachment=True,
        download_name=f"재고현황_{오늘}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp_재고현황.route("/api/재고현황/변경확인", methods=["GET"])
def 재고현황_변경확인():
    """
    폴링용. 클라이언트가 마지막으로 본 타임스탬프 전달 → 변경 여부 반환.
    ?ts=2026-04-17 10:00:00
    변경됐으면 rows 전체 반환, 아니면 changed: false
    """
    클라_ts = request.args.get("ts", "")
    저장_파일 = config.재고현황_저장_파일
    if not 저장_파일.exists():
        return jsonify({"changed": False})

    try:
        with open(저장_파일, encoding="utf-8") as f:
            저장_data = json.load(f)
        서버_ts = 저장_data.get("last_modified", "")
        if 서버_ts == 클라_ts:
            return jsonify({"changed": False})
        return jsonify({
            "changed": True,
            "last_modified": 서버_ts,
            "rows": 저장_data.get("rows", []),
        })
    except Exception as e:
        return jsonify({"changed": False})


@bp_재고현황.route("/api/재고현황/분석", methods=["POST"])
def 재고현황_분석():
    """
    multipart/form-data:
      재고현황표, 사방넷, basic, forecast, 누아트sku, po(복수), 주문서
      제외_날짜들: JSON 문자열 (선택)
    """
    try:
        # 임시 폴더 생성
        temp_id = uuid.uuid4().hex[:8]
        temp_dir = config.재고현황_UPLOAD_DIR / temp_id
        temp_dir.mkdir(parents=True, exist_ok=True)

        def _안전파일명(원본: str, fallback: str) -> str:
            """한글 등 non-ASCII 파일명도 안전하게 처리."""
            ext = Path(원본).suffix or ""
            name = secure_filename(원본)
            if not name or name == ext.lstrip("."):
                name = fallback + ext
            return name

        def _저장(field: str) -> Path | None:
            f = request.files.get(field)
            if not f or not f.filename:
                return None
            name = _안전파일명(f.filename, field)
            경로 = temp_dir / name
            f.save(str(경로))
            return 경로

        def _저장_복수(field: str) -> list[Path]:
            files = request.files.getlist(field)
            결과 = []
            for f in files:
                if f and f.filename:
                    name = _안전파일명(f.filename, field)
                    경로 = temp_dir / name
                    f.save(str(경로))
                    결과.append(경로)
            return 결과

        재고현황표 = _저장("재고현황표")   # 선택 사항
        사방넷      = _저장("사방넷")
        basic       = _저장("basic")
        forecast    = _저장("forecast")
        누아트sku   = _저장("누아트sku")
        po_파일들   = _저장_복수("po")
        주문서      = _저장("주문서")

        제외_날짜들 = json.loads(request.form.get("제외_날짜들", "[]"))

        # 기존 저장 데이터 로드 → 부분 업로드 시 없는 파일의 컬럼 값 유지
        기존_행맵: dict = {}
        if config.재고현황_저장_파일.exists():
            try:
                with open(config.재고현황_저장_파일, encoding="utf-8") as f:
                    기존 = json.load(f)
                기존_행맵 = {r["바코드"]: r for r in 기존.get("rows", []) if r.get("바코드")}
            except Exception:
                pass

        uc = 재고현황_유스케이스(sheets_저장소=_sheets_연결())
        결과 = uc.분석(
            사방넷_파일=사방넷,
            basic_파일=basic,
            forecast_파일=forecast,
            누아트sku_파일=누아트sku,
            po_파일들=po_파일들,
            주문서_파일=주문서,
            재고현황표_파일=재고현황표,
            제외_날짜들=제외_날짜들,
            기존_행맵=기존_행맵,
        )

        shutil.rmtree(temp_dir, ignore_errors=True)

        # 분석 성공 시 결과 저장 (업데이트 시각 포함)
        if 결과.get("success"):
            결과["업데이트_시각"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            config.재고현황_저장_파일.parent.mkdir(parents=True, exist_ok=True)
            with open(config.재고현황_저장_파일, "w", encoding="utf-8") as f:
                json.dump(결과, f, ensure_ascii=False)
            _캐시_갱신(결과)
            _아카이브_저장(결과)
            # 구글 시트 스냅샷 누적 (같은 날 재실행 시 덮어씀)
            try:
                sheets = _sheets_연결()
                if sheets:
                    sheets.스냅샷_저장(결과.get("rows", []))
            except Exception as _e:
                print(f"[스냅샷] 구글시트 저장 실패 (분석 결과에는 영향 없음): {_e}")

        return jsonify(결과)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
