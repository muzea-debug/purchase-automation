"""
발주 실행 라우트 - 바코드 조회 / 발주 실행 API
"""
import io
from flask import Blueprint, render_template, request, jsonify, send_file, abort
import config
from application.발주_실행_유스케이스 import 발주_실행_유스케이스
bp_발주 = Blueprint("발주", __name__)

try:
    _유스케이스 = 발주_실행_유스케이스()
    _파서 = _유스케이스._파서
except Exception as _init_err:
    import traceback
    print(f"[routes_발주] 초기화 실패: {_init_err}\n{traceback.format_exc()}", flush=True)
    _유스케이스 = None
    _파서 = None


@bp_발주.route("/발주")
def 발주_페이지():
    return render_template("발주_실행.html")


@bp_발주.route("/api/발주/카탈로그", methods=["GET"])
def 카탈로그_현황():
    """통합발주서원본/ 폴더 스캔 결과 반환."""
    파일_목록 = sorted(f for f in config.통합발주서원본_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
    카탈로그 = _파서.카탈로그_로드()

    # 파일별 상품 수 집계
    파일별 = {}
    for 상품 in 카탈로그.values():
        key = 상품.원본파일.name
        파일별[key] = 파일별.get(key, 0) + 1

    files = [
        {"파일명": f.name, "상품수": 파일별.get(f.name, 0)}
        for f in 파일_목록
    ]
    return jsonify({
        "success": True,
        "총_상품수": len(카탈로그),
        "파일수": len(파일_목록),
        "files": files,
    })


@bp_발주.route("/api/발주/상품검색", methods=["GET"])
def 상품검색():
    """제품명/바코드 자동완성용. q= 검색어, 최대 20개 반환."""
    if _파서 is None:
        return jsonify({"items": []})
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"items": []})
    카탈로그 = _파서.카탈로그_로드()
    q_lower = q.lower()
    results = []
    for bc, info in 카탈로그.items():
        if q_lower in info.이름.lower() or q in bc or q_lower in info.브랜드.lower():
            results.append({"바코드": bc, "이름": info.이름, "브랜드": info.브랜드})
            if len(results) >= 20:
                break
    return jsonify({"items": results})


@bp_발주.route("/api/발주/조회", methods=["POST"])
def 발주_조회():
    """
    body: {"items": [{"barcode": "...", "quantity": 500}, ...]}
    """
    if _유스케이스 is None:
        return jsonify({"success": False, "error": "서버 초기화 실패 - 콘솔 로그 확인"}), 500
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        if not items:
            return jsonify({"success": False, "error": "items가 비어있습니다."}), 400

        result = _유스케이스.조회(items)
        return jsonify(result)
    except Exception as e:
        import traceback
        print(f"[발주_조회 오류] {e}\n{traceback.format_exc()}", flush=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp_발주.route("/api/발주/카탈로그_상태", methods=["GET"])
def 카탈로그_상태():
    """캐시 상태 + 디렉토리 존재 여부 확인용."""
    import config as _cfg
    dir_exists = _cfg.통합발주서원본_DIR.exists()
    files = []
    if dir_exists:
        files = [f.name for f in _cfg.통합발주서원본_DIR.glob("*.xlsx")
                 if not f.name.startswith("~$")]
    캐시_크기 = len(_파서._캐시)
    return jsonify({
        "디렉토리": str(_cfg.통합발주서원본_DIR),
        "디렉토리_존재": dir_exists,
        "xlsx_파일수": len(files),
        "캐시_상품수": 캐시_크기,
    })


@bp_발주.route("/api/발주/실행", methods=["POST"])
def 발주_실행():
    """
    body:
      발주_유형: "정기발주" | "신기종"
      한국_요청_도착일: "2026. 4. 30"   (optional)
      groups: [{원본파일, po코드, items: [{barcode, quantity}]}]
    """
    payload = request.get_json(force=True)
    if not payload.get("groups"):
        return jsonify({"success": False, "error": "groups가 비어있습니다."}), 400

    result = _유스케이스.실행(payload)
    status = 200 if result.get("success") else 400
    return jsonify(result), status


@bp_발주.route("/api/발주/이메일_발송", methods=["POST"])
def 이메일_발송():
    """발주 완료 후 수동으로 이메일 발송."""
    if _유스케이스 is None:
        return jsonify({"success": False, "error": "서버 초기화 실패"}), 500
    data = request.get_json(force=True)
    이메일_준비 = data.get("이메일_준비", {})
    결과 = _유스케이스.이메일_발송_직접(
        파일_경로들=이메일_준비.get("파일_경로들", []),
        po코드들=이메일_준비.get("po코드들", []),
        사유=이메일_준비.get("사유", "정기발주"),
        총수량=이메일_준비.get("총수량", 0),
    )
    if 결과 == "":
        return jsonify({"success": True, "message": "이메일 발송 완료"})
    elif 결과 == "미설정":
        return jsonify({"success": False, "error": "이메일 설정이 없습니다 (.env SMTP 확인)"}), 400
    else:
        return jsonify({"success": False, "error": 결과}), 500


@bp_발주.route("/api/발주/이메일_설정", methods=["GET"])
def 이메일_설정_조회():
    """저장된 이메일 수신자 설정 반환. 없으면 env 값 반환."""
    import json
    설정 = {"수신자": config.SMTP_수신자, "참조": config.SMTP_참조}
    if config.이메일_설정_파일.exists():
        try:
            저장 = json.loads(config.이메일_설정_파일.read_text(encoding="utf-8"))
            설정.update(저장)
        except Exception:
            pass
    return jsonify({"success": True, **설정})


@bp_발주.route("/api/발주/이메일_설정", methods=["POST"])
def 이메일_설정_저장():
    """이메일 수신자 설정 저장. body: {"수신자": "...", "참조": "..."}"""
    import json
    data = request.get_json(force=True)
    수신자 = str(data.get("수신자") or "").strip()
    참조   = str(data.get("참조") or "").strip()
    if not 수신자:
        return jsonify({"success": False, "error": "수신자 이메일을 입력해주세요."}), 400
    설정 = {"수신자": 수신자, "참조": 참조}
    config.이메일_설정_파일.parent.mkdir(parents=True, exist_ok=True)
    config.이메일_설정_파일.write_text(json.dumps(설정, ensure_ascii=False), encoding="utf-8")
    return jsonify({"success": True})


@bp_발주.route("/api/발주/발주현황_헤더확인", methods=["GET"])
def 발주현황_헤더확인():
    """4월 발주현황 시트 헤더를 읽어 현재 매핑 결과 반환."""
    if _유스케이스 is None:
        return jsonify({"success": False, "error": "서버 초기화 실패"}), 500
    sheets = _유스케이스._sheets
    if not sheets or not hasattr(sheets, '_발주현황_컬럼_매핑'):
        return jsonify({"success": False, "error": "Google Sheets 미연결"}), 400
    from datetime import datetime as _dt
    탭명 = f"{_dt.now().month}월 발주현황"
    try:
        result = sheets._service.spreadsheets().values().get(
            spreadsheetId=sheets._스프레드시트_id,
            range=f"'{탭명}'!1:1",
        ).execute()
        헤더_행 = result.get("values", [[]])[0] if result.get("values") else []
        헤더_lower = [str(h).strip().lower() for h in 헤더_행]
        col_맵 = sheets._발주현황_컬럼_매핑(헤더_lower)
        # 열 인덱스 → 엑셀 열 문자 변환
        def col_letter(i): return chr(ord('A') + i) if i < 26 else chr(ord('A') + i // 26 - 1) + chr(ord('A') + i % 26)
        매핑_결과 = {field: f"{col_letter(idx)}열 ({헤더_행[idx] if idx < len(헤더_행) else '새열'})" for field, idx in sorted(col_맵.items(), key=lambda x: x[1])}
        미매핑 = [f for f in col_맵 if col_맵[f] >= len(헤더_행)]
        return jsonify({
            "success": True,
            "탭명": 탭명,
            "헤더": 헤더_행,
            "매핑": 매핑_결과,
            "미매핑_필드": 미매핑,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp_발주.route("/api/발주/양식_다운로드", methods=["GET"])
def 양식_다운로드():
    """발주 입력 양식 xlsx 반환 (바코드 / 수량 / 발주유형)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주입력"

    헤더 = ["바코드", "수량", "발주유형"]
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(헤더, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 예시 행
    ws.append(["8809961834097", 100, "정기발주"])
    ws.append(["8809961835339", 50, "정기발주"])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="발주입력_양식.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp_발주.route("/api/발주/엑셀_파싱", methods=["POST"])
def 엑셀_파싱():
    """업로드된 xlsx를 파싱해 items 리스트 반환."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"success": False, "error": "파일이 없습니다."}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"success": False, "error": f"파일 읽기 실패: {e}"}), 400

    if not rows:
        return jsonify({"success": False, "error": "빈 파일입니다."}), 400

    # 헤더 자동 탐지 (바코드, 수량, 발주유형 컬럼 위치)
    헤더_행 = [str(v or "").strip().lower() for v in rows[0]]
    def _col(*keywords):
        for k in keywords:
            for i, h in enumerate(헤더_행):
                if k in h:
                    return i
        return None

    bc_col  = _col("바코드", "barcode")
    qty_col = _col("수량", "qty", "quantity")
    사유_col = _col("발주유형", "사유", "type")

    if bc_col is None or qty_col is None:
        return jsonify({"success": False, "error": "바코드/수량 컬럼을 찾을 수 없습니다."}), 400

    items = []
    errors = []
    for i, row in enumerate(rows[1:], 2):
        bc = str(row[bc_col] or "").strip() if bc_col < len(row) else ""
        qty_raw = row[qty_col] if qty_col < len(row) else None
        사유 = str(row[사유_col] or "정기발주").strip() if 사유_col is not None and 사유_col < len(row) else "정기발주"

        if not bc:
            continue
        try:
            qty = int(float(str(qty_raw or 0)))
        except (ValueError, TypeError):
            errors.append(f"{i}행: 수량 오류 ({qty_raw})")
            continue
        if qty <= 0:
            errors.append(f"{i}행: 수량이 0 이하 ({bc})")
            continue
        items.append({"barcode": bc, "quantity": qty, "사유": 사유})

    # 제품명 조회 (카탈로그에서)
    if _파서 and items:
        try:
            카탈로그 = _파서.카탈로그_로드()
            for it in items:
                상품 = 카탈로그.get(it["barcode"])
                it["이름"] = 상품.이름 if 상품 else ""
        except Exception:
            pass

    return jsonify({"success": True, "items": items, "errors": errors, "총행수": len(items)})


@bp_발주.route("/api/발주/po_대시보드", methods=["GET"])
def po_대시보드():
    """각 PO 시리즈의 현재 활성 prefix + 최대번호 반환."""
    import string as _string
    from domain.services.발주_코드_서비스 import 발주_코드_서비스

    결과 = []
    for 시리즈 in config.PO_시리즈:
        base = 시리즈["base"]
        name = 시리즈["name"]

        def _max_fn(prefix, _base=base):
            if _유스케이스 and _유스케이스._sheets:
                try:
                    n = _유스케이스._sheets.마지막_po번호_조회(prefix)
                    if n > 0:
                        return n
                except Exception:
                    pass
            return 발주_코드_서비스.최대번호_로컬(config.발주서_DIR, prefix)

        try:
            prefix, current = 발주_코드_서비스.현재_활성_prefix(base, _max_fn, config.PO_번호_상한)
            결과.append({
                "base": base,
                "name": name,
                "prefix": prefix,
                "current": current,
                "max": config.PO_번호_상한,
                "full": current >= config.PO_번호_상한,
            })
        except Exception as e:
            결과.append({"base": base, "name": name, "prefix": base + "A",
                         "current": 0, "max": config.PO_번호_상한, "full": False})

    return jsonify({"success": True, "series": 결과})


@bp_발주.route("/api/발주/po_대시보드", methods=["POST"])
def po_대시보드_수정():
    """대시보드에서 수동으로 특정 prefix의 현재 번호를 오버라이드."""
    import json as _json
    data = request.get_json(force=True)
    prefix = str(data.get("prefix", "")).strip()
    번호 = int(data.get("current", 0))
    if not prefix:
        return jsonify({"success": False, "error": "prefix 없음"}), 400

    오버라이드_파일 = config.발주서_DIR / "po_시작번호.json"
    try:
        오버라이드_파일.parent.mkdir(parents=True, exist_ok=True)
        기존 = {}
        if 오버라이드_파일.exists():
            기존 = _json.loads(오버라이드_파일.read_text(encoding="utf-8"))
        기존[prefix] = 번호
        오버라이드_파일.write_text(_json.dumps(기존, ensure_ascii=False, indent=2), encoding="utf-8")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp_발주.route("/api/발주/po_시리즈_추가", methods=["POST"])
def po_시리즈_추가():
    """
    신규 PO 유형(시리즈) 추가.
    body: {"base": "NCA", "name": "누아트ACC", "category": "브랜드"}
    config.PO_시리즈 목록에 런타임으로 추가하고, 영속화용 JSON 파일에도 저장.
    """
    import json as _json
    data = request.get_json(force=True)
    base = str(data.get("base", "")).strip().upper()
    name = str(data.get("name", "")).strip()
    category = str(data.get("category", "브랜드")).strip()

    if not base or not name:
        return jsonify({"success": False, "error": "base와 name은 필수입니다"}), 400

    # 중복 확인
    if any(s["base"] == base for s in config.PO_시리즈):
        return jsonify({"success": False, "error": f"이미 존재하는 코드입니다: {base}"}), 400

    새_시리즈 = {"base": base, "name": name, "category": category}
    config.PO_시리즈.append(새_시리즈)

    # BRAND_PO_BASE에도 추가 (발주 코드 자동생성용)
    if category == "부자재":
        config.BRAND_전장품_PO_BASE[name] = base
    else:
        config.BRAND_PO_BASE[name] = base

    # 영속화: 추가 시리즈 JSON 파일
    추가_시리즈_파일 = config.발주서_DIR / "po_추가시리즈.json"
    try:
        추가_시리즈_파일.parent.mkdir(parents=True, exist_ok=True)
        기존 = []
        if 추가_시리즈_파일.exists():
            기존 = _json.loads(추가_시리즈_파일.read_text(encoding="utf-8"))
        기존.append(새_시리즈)
        추가_시리즈_파일.write_text(_json.dumps(기존, ensure_ascii=False, indent=2), encoding="utf-8")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp_발주.route("/api/발주/신규알림", methods=["GET"])
def 신규알림():
    """원본 발주서 생성 시 발주 실행 페이지로 전달할 알림 반환 (읽으면 삭제)."""
    try:
        from application.알림_저장소 import 알림_읽기_후_삭제
        items = 알림_읽기_후_삭제()
        return jsonify({"success": True, "items": items})
    except Exception as e:
        return jsonify({"success": True, "items": []})


@bp_발주.route("/발주/다운로드/<order_type>/<filename>")
def 발주서_다운로드(order_type, filename):
    """생성된 발주서 xlsx 다운로드."""
    if "/" in filename or ".." in filename or "/" in order_type or ".." in order_type:
        abort(400)
    경로 = config.발주서_DIR / order_type / filename
    if not 경로.exists():
        abort(404)
    return send_file(
        경로,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
